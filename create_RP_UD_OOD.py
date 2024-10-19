"""
Скрипт для отработки генерации рабочих программ учебных дисциплин общеобразовательного типа с помощью шаблонов docxtemplate
"""
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from docxtpl import DocxTemplate
import string
import time
import re
from tkinter import messagebox
from jinja2 import exceptions

pd.options.mode.chained_assignment = None  # default='warn'
pd.set_option('display.max_columns', None)  # Отображать все столбцы
pd.set_option('display.expand_frame_repr', False)  # Не переносить строки
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.filterwarnings('ignore', category=FutureWarning, module='openpyxl')

class DiffSheet(Exception):
    """
    Исключение для случаев когда отсутствуют нужные колонки в файле
    """
    pass

class ControlSemestr(Exception):
    """
    Исключение для случаев когда на листе План Мдк отсутствует слово семест в колонках
    """
    pass

class ControlChapter(Exception):
    """
    Исключение для обработки случае если на листе План УД отсутствуют записи со словом Раздел
    """
    pass

class NotDataMdk(Exception):
    """
    Исключения для случаев когда пустой датафрейм в плане МДК
    """
    pass

def convert_to_int(cell):
    """
    Метод для проверки значения ячейки
    :param cell: значение ячейки
    :return: число в формате int
    """
    if cell is np.nan:
        return ''
    try:
        value = float(cell)
        return int(value)
    except:
        return cell


def processing_punctuation_end_string(lst_phrase: list, sep_string: str, sep_begin: str, sep_end: str) -> str:
    """
    Очистка каждого элемента списка от знаков пунктации в конце,
    добавление разделитея, добавление точки в конце
    :param lst_phrase: список элементов
    :param sep_string: разделитель между элементами списка
    :param sep_begin: начальный разделитель
    :param sep_end: знак пунктуации в конце
    :return: строку с разделителями и переносами строки
    """
    if len(lst_phrase) == 0:
        return ''
    lst_phrase = list(map(str,lst_phrase)) # делаем строковыми элементы
    temp_lst = list(map(lambda x: sep_begin + x, lst_phrase))
    temp_lst = list(map(lambda x: x.strip(), temp_lst)) # очищаем от прбельных символов в начале и конце
    temp_lst = list(map(lambda x: x.rstrip(string.punctuation), temp_lst))  # очищаем от знаков пунктуации
    temp_lst[-1] = temp_lst[-1] + sep_end  # добавляем конечный знак пунктуации
    temp_str = f'{sep_string}'.join(temp_lst)  # создаем строку с разделителями
    return temp_str

def insert_type_source(df:pd.DataFrame)->list:
    """
    Вставка в строку слов [Электронный ресурс] Форма доступа:
    :param lst_phrase:датафрейм
    :return: список измененных строк
    """
    out_lst = [] # список для хранения строк
    for row in df.itertuples():
        name = str(row[1])
        url_ii = str(row[2])
        name = name.strip() # очищаем от пробельных символов
        name = name.rstrip(string.punctuation) # очищаем от знаков препинания
        name = name.strip()
        url_ii = url_ii.strip() # очищаем от пробельных символов
        url_ii = url_ii.rstrip(string.punctuation) # очищаем от знаков препинания
        url_ii = url_ii.strip()
        temp_str = f'{name} [Электронный ресурс] Форма доступа:{url_ii}'
        out_lst.append(temp_str)

    return out_lst

def processing_publ(row):
    """
    Функция для генерации строки с литературой в нужном формате
    :param row: строка датафрейма
    :return: сумма строк датафрейма в нужном формате
    """
    author = row[0] # автор(ы)
    name_book = row[1] # название
    full_city = row[2] # полное название города
    short_city = row[3] # краткое название города
    publ_house = row[4] # издательство
    year = row[5] # год издания
    quan_pages = row[6] # число страниц
    author = author.rstrip(string.punctuation) # очищаем от символа пунктуации в конце
    name_book = name_book.rstrip(string.punctuation) # очищаем от символа пунктуации в конце
    short_city = short_city.rstrip(string.punctuation)  # очищаем от символа пунктуации в конце
    publ_house = publ_house.rstrip(string.punctuation)  # очищаем от символа пунктуации в конце
    # извлекаем год
    result = re.search(r'\d{4}',year)
    if result:
        clean_year = result.group()
    else:
        clean_year = 'Неправильно заполнен год издания, введите год в формате 4 цифры без букв'

    # извлекаем количество страниц
    result = re.search(r'\d+',quan_pages)
    if result:
        clean_quan_pages = result.group()
    else:
        clean_quan_pages = 'Неправильно заполнено количество страниц, введите количество в виде числа без букв'
    # Формируем итоговую строку
    out_str = f'{author}. {name_book}.- {short_city}.: {publ_house}, {clean_year}.- {clean_quan_pages} c.'

    return out_str

def write_df_to_excel(dct_df:dict,write_index:bool)->openpyxl.Workbook:
    """
    Функция для записи датафрейма в файл Excel
    :param dct_df: словарь где ключе это название создаваемого листа а значение датафрейм который нужно записать
    :param write_index: нужно ли записывать индекс датафрейма True or False
    :return: объект Workbook с записанными датафреймами
    """
    wb = openpyxl.Workbook() # создаем файл
    count_index = 0 # счетчик индексов создаваемых листов
    for name_sheet,df in dct_df.items():
        wb.create_sheet(title=name_sheet,index=count_index) # создаем лист
        # записываем данные в лист
        for row in dataframe_to_rows(df,index=write_index,header=True):
            wb[name_sheet].append(row)
        # ширина по содержимому
        # сохраняем по ширине колонок
        for column in wb[name_sheet].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[name_sheet].column_dimensions[column_name].width = adjusted_width
        count_index += 1
    # удаляем лишний лист
    if len(wb.sheetnames) >= 2 and 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    return wb


def sum_column_any_value(df:pd.DataFrame,name_column:str):
    """
    Суммирование колонки с разными типами значений в том числе строковыми.
    """
    lst_value = df[name_column].dropna().tolist()
    sum_value = [value for value in lst_value if isinstance(value,(int,float))] # отбираем только числа
    return sum(sum_value) # возвращаем сумму

def create_check_error_df(dct:dict)->pd.DataFrame:
    """
    Функция для разворачивания словаря с данными по каждому мдк в датафрейм
    :param dct: Словарь с данными
    :type dct:dict
    :return:Датафрейм
    :rtype:pd.Dataframe
    """
    df = pd.DataFrame(columns=['Семестр','Практическая подготовка','Обязательная нагрузка',
                               'Прак_лаб занятия','КР','Урок','Практическое занятие','Лабораторное занятие'])
    for name_sem,part in dct.items():
        prac_hour = part['практическое занятие'] + part['лабораторное занятие'] # считаем практические занятия
        all_hours = part['практическое занятие'] + part['лабораторное занятие'] + part['урок'] + part['курсовая работа (КП)']
        # создаем строку датафрейма
        temp_df = pd.DataFrame(columns=['Семестр','Практическая подготовка','Обязательная нагрузка',
                           'Прак_лаб занятия','КР','Урок','Практическое занятие','Лабораторное занятие'],
                               data=[[name_sem,part['Прак_подготовка'],all_hours,prac_hour,part['курсовая работа (КП)'],
                                      part['урок'],part['практическое занятие'],part['лабораторное занятие']]])
        df = pd.concat([df,temp_df],ignore_index=True,axis=0)

    sum_row = df.sum()  # получаем строку общей суммы
    df.loc['Сумма'] = sum_row  # добавляем строку в датафрейм
    df.at['Сумма', 'Семестр'] = 'Итого'
    return df

def add_competence_column(df:pd.DataFrame,dct_competence:dict):
    """
    Функция для добавления колонки в которой для каждой темы будет указаны используемые ОК и ПК
    :param df: датафрейм
    :param dct_competence:  словарь с данными в каких разделах и темах используются ОК и ПК
    :return: датафрейм с колонкой для ОК и ПК
    """
    check_chapter = None

    for row in df.itertuples():
        row_content = row[2]
        if 'Раздел' in row_content:
            result_chapter = re.search(r'Раздел\s*\d+',row_content) # Ищем сочетание Раздел цифра
            if result_chapter:
                check_chapter = result_chapter.group(0) # указываем текущий раздел
                if check_chapter in dct_competence:
                    chapter_str_competence_lst = [] # список для ОК и ПК используемых в разделе
                    for key,value in dct_competence[check_chapter].items():
                        for comp in value:
                            # Извлекаем номера ОК и ПК
                            comp_result = re.search(r'ОК\s*\d+|ПК\s+\d+\.?\d+',comp)
                            if comp_result:
                                chapter_str_competence_lst.append(comp_result.group(0))
                    chapter_str_competence_lst.sort() # сортируем чтобы ОК были на первом месте
                    df.iloc[row[0],6]=','.join(chapter_str_competence_lst) # записываем результат в колонку ОК_ПК
        elif 'Тема' in row_content:
            if check_chapter: # делаем обработку только если какой-то раздел был уже найден
                result_theme = re.search(r'\b\d+\.\d+\b',row_content)
                if result_theme:
                    theme = result_theme.group(0)
                    if theme in dct_competence[check_chapter]:
                        # Извлекаем ОК и ПК
                        chapter_str_competence_lst = []  # список для ОК и ПК используемых в разделе
                        for value in dct_competence[check_chapter][theme]:
                            # Извлекаем номера ОК и ПК
                            comp_result = re.search(r'ОК\s*\d+|ПК\s+\d+\.?\d+', value)
                            if comp_result:
                                chapter_str_competence_lst.append(comp_result.group(0))
                        chapter_str_competence_lst.sort() # сортируем чтобы ОК были на первом месте
                        df.iloc[row[0],6]=','.join(chapter_str_competence_lst) # записываем результат в колонку ОК_ПК
                    else:
                        if check_chapter in dct_competence:
                            chapter_str_competence_lst = []  # список для ОК и ПК используемых в разделе
                            for value in dct_competence[check_chapter][check_chapter]:
                                # Извлекаем номера ОК и ПК
                                comp_result = re.search(r'ОК\s*\d+|ПК\s+\d+\.?\d+', value)
                                if comp_result:
                                    chapter_str_competence_lst.append(comp_result.group(0))
                            chapter_str_competence_lst.sort()  # сортируем чтобы ОК были на первом месте
                            df.iloc[row[0], 6] = ','.join(
                                chapter_str_competence_lst)  # записываем результат в колонку ОК_ПК

    return df






def extract_data_plan_ood(data_ud, sheet_name,dct_competence:dict):
    """
    Функция для получения датафрейма из листа файла
    :param data_ud: путь к файлу
    :param sheet_name: имя листа
    :param dct_competence: словарь где указано в каких разделах и темах содержится то или иное ОК, ПК
    :return: датафрейм
    """
    print(sheet_name)
    lst_type_lesson = ['урок', 'практическое занятие', 'лабораторное занятие',
                       'курсовая работа (КП)']  # список типов занятий
    dct_all_sum_result = {key: 0 for key in lst_type_lesson}  # создаем словарь для подсчета значений


    # Создаем датафрейм
    df_plan_pm = pd.read_excel(data_ud,sheet_name=sheet_name, usecols='A:H')
    if df_plan_pm.shape[0] == 0:
        raise NotDataMdk
    df_plan_pm.dropna(inplace=True, thresh=1)  # удаляем пустые строки

    # Устанавливаем стандартные колонки
    df_plan_pm.columns = ['Курс_семестр', 'Раздел', 'Тема', 'Содержание', 'Количество_часов', 'Прак_подготовка', 'Вид_занятия',
                          'СРС']
    df_plan_pm['Курс_семестр'].fillna('Пусто', inplace=True)
    df_plan_pm['Раздел'].fillna('Пусто', inplace=True)
    df_plan_pm['Тема'].fillna('Пусто', inplace=True)

    # Считаем общие суммы
    mdk_all_sum = int(sum_column_any_value(df_plan_pm, 'Количество_часов'))  # получаем сумму общие часы
    mdk_all_prac_sum = int(sum_column_any_value(df_plan_pm, 'Прак_подготовка'))  # получаем сумму общие часы
    mdk_all_srs_sum = int(sum_column_any_value(df_plan_pm, 'СРС'))  # сумма срс
    for type_lesson in lst_type_lesson:
        _df = df_plan_pm[df_plan_pm['Вид_занятия'] == type_lesson]  # фильтруем датафрейм
        dct_all_sum_result[type_lesson] = int(sum_column_any_value(_df, 'Количество_часов'))  # получаем значение

    dct_all_sum_result['Всего часов'] = mdk_all_sum
    dct_all_sum_result['Всего прак_подготовка'] = mdk_all_prac_sum
    dct_all_sum_result['Всего СРС'] = mdk_all_srs_sum


    borders = df_plan_pm[
        df_plan_pm['Курс_семестр'].str.contains('семестр')].index  # получаем индексы строк где есть слово семестр

    if len(borders) == 0:
        raise ControlSemestr # если не находим ни одной ячейки где есть слово семестр то вызываем исключение

    name_borders = [] # лист для хранения названий семестров(границ разделов)
    for idx in borders:
        _temp_name = df_plan_pm.at[idx,'Курс_семестр'].replace('\n','') # получаем название семестров и удаляем символ переноса
        name_borders.append(_temp_name)

    part_df = []  # список для хранения кусков датафрейма
    previos_border = -1
    # делим датафрем по границам
    for value_border in borders:
        part = df_plan_pm.iloc[previos_border:value_border]
        part_df.append(part)
        previos_border = value_border

    # добавляем последнюю часть
    last_part = df_plan_pm.iloc[borders[-1]:]
    part_df.append(last_part)

    part_df.pop(0)  # удаляем нулевой элемент так как он пустой

    main_df = pd.DataFrame(
        columns=['Курс_семестр', 'Раздел', 'Тема', 'Содержание', 'Количество_часов', 'Прак_подготовка', 'Вид_занятия',
                 'СРС'])  # создаем базовый датафрейм
    part_dct_sum = dict() # создаем словарь который будет хранить суммы по семестрам

    for idx,df in enumerate(part_df):
        dct_sum_result = {key: 0 for key in lst_type_lesson}  # создаем словарь для подсчета значений
        dct_sum_result['Прак_подготовка'] = 0
        for type_lesson in lst_type_lesson:
            _df = df[df['Вид_занятия'] == type_lesson]  # фильтруем датафрейм
            _df['Количество_часов'].fillna(0, inplace=True)
            dct_sum_result[type_lesson] = int(sum_column_any_value(_df,'Количество_часов')) # считаем часы для каждого типа
            dct_sum_result['Прак_подготовка'] += int(sum_column_any_value(_df,'Прак_подготовка')) # считаем часы для практики
        part_dct_sum[name_borders[idx]] = dct_sum_result

        # создаем строку с описанием
        margint_text = 'Итого часов за семестр:\nиз них\nтеория\nпрактические занятия\nлабораторные занятия\nкурсовая работа (КП)'

        not_prac_dct_sum_result = dct_sum_result.copy()  # получаем копию
        not_prac_dct_sum_result.pop('Прак_подготовка')  # удаляем ключ практическая подготовка
        all_hours = sum(not_prac_dct_sum_result.values())  # общая сумма часов

        theory_hours = dct_sum_result['урок']  # часы теории
        praktice_hours = dct_sum_result['практическое занятие']  # часы практики
        lab_hours = dct_sum_result['лабораторное занятие']  # часы лабораторных
        kurs_hours = dct_sum_result['курсовая работа (КП)']  # часы курсовых

        value_text = f'{all_hours}\n \n{theory_hours}\n{praktice_hours}\n{lab_hours}\n{kurs_hours}'  # строка со значениями
        temp_df = pd.DataFrame([{'Тема': margint_text, 'Количество_часов': value_text}])
        df = pd.concat([df, temp_df], ignore_index=True)  # добаляем итоговую строку
        main_df = pd.concat([main_df, df], ignore_index=True)  # добавляем в основной датафрейм

    main_df.insert(0, 'Номер', np.nan)  # добавляем колонку с номерами занятий

    main_df['Содержание'] = main_df['Содержание'].fillna('Пусто')  # заменяем наны на пусто

    count = 0  # счетчик
    for idx, row in enumerate(main_df.itertuples()):
        if (row[5] == 'Пусто') | ('Итого часов' in row[5]):
            main_df.iloc[idx, 0] = ''
        else:
            count += 1
            main_df.iloc[idx, 0] = count

    # очищаем от пустых символов и строки Пусто
    main_df['Курс_семестр'] = main_df['Курс_семестр'].fillna('Пусто')
    main_df['Раздел'] = main_df['Раздел'].fillna('Пусто')

    main_df['Курс_семестр'] = main_df['Курс_семестр'].replace('Пусто', '')
    main_df['Тема'] = main_df['Тема'].replace('Пусто', '')
    main_df['Раздел'] = main_df['Раздел'].replace('Пусто', '')
    main_df['Содержание'] = main_df['Содержание'].replace('Пусто', '')

    main_df['Вид_занятия'] = main_df['Вид_занятия'].fillna('')

    main_df['Количество_часов'] = main_df['Количество_часов'].apply(convert_to_int)
    main_df['Количество_часов'] = main_df['Количество_часов'].fillna('')
    main_df['Прак_подготовка'] = main_df['Прак_подготовка'].fillna(0)
    main_df['Прак_подготовка'] = main_df['Прак_подготовка'].astype(int, errors='ignore')
    main_df['Прак_подготовка'] = main_df['Прак_подготовка'].apply(lambda x: '' if x == 0 else x)

    main_df['СРС'] = main_df['СРС'].fillna(0)
    main_df['СРС'] = main_df['СРС'].astype(int, errors='ignore')
    main_df['СРС'] = main_df['СРС'].apply(lambda x: '' if x == 0 else x)
    main_df['Содержание'] = main_df['Курс_семестр'] + main_df['Раздел'] + main_df['Тема'] + main_df['Содержание']
    main_df.drop(columns=['Курс_семестр', 'Раздел', 'Тема'], inplace=True)

    # Добавляем колонку с ОК и ПК
    main_df['ОК_ПК'] = ''
    main_df = add_competence_column(main_df,dct_competence)

    return (main_df,dct_all_sum_result,part_dct_sum) # возвращаем кортеж

def find_part_themes(value:str,dct_competence:dict,competence:str):
    """
    Функция для извлечения из ячеек в колонке Раздел_тема разделов и тем где используется то или иное ОК или ПК
    :param value: строка
    :param dct_competence: словарь где будут храниться данные
    :param competence: наименование ПК или ОК
    :return: словарь
    """

    lst_part = value.split(',')
    lst_part = [value for value in lst_part if value] # очищаем от последнего элемента на случай если в конце стоит запятая
    for part in lst_part:
        temp_part = part.split(':')  # извлекаем название раздела и часть с темами
        if len(temp_part) != 2:  # если отсутствует двоеточиее не обрабатываем
            if len(temp_part) == 1:
                result_part = re.search(r'Раздел\s*\d+', temp_part[0])  # Выделяем название раздела
                if result_part:
                    name_chapter = result_part.group(0) # название раздела
                    # Проверяем наличие такого ключа в словаре
                    if name_chapter not in dct_competence:
                        dct_competence[name_chapter] = {}  # если нет то создаем такой ключ где значением будет словарь
                    # Проверяем есть ключ с названием раздела внутри
                    if name_chapter not in dct_competence[name_chapter]:
                        dct_competence[name_chapter][name_chapter] = [competence]
                    else:
                        dct_competence[name_chapter][name_chapter].append(competence)


        else:
            temp_part = list(map(str.strip, temp_part))  # очищаем от пробелов в начале и конце
            result_part = re.search(r'Раздел\s*\d+', temp_part[0])  # Выделяем название раздела
            if result_part:
                name_chapter = result_part.group(0) # название раздела
                # Проверяем наличие такого ключа в словаре
                if name_chapter not in dct_competence:
                    dct_competence[name_chapter] = {} # если нет то создаем такой ключ где значением будет словарь

                lst_result_themes = re.findall(r'\b\d+\.\d+\b', temp_part[1])
                if len(lst_result_themes) == 0:
                    if name_chapter in dct_competence[name_chapter]:
                        dct_competence[name_chapter][name_chapter].append(competence) # создаем внутренний ключ с названием раздела
                    else:
                        dct_competence[name_chapter] = {name_chapter:[competence]}
                else:
                    # перебираем найденные темы
                    for theme in lst_result_themes:
                        if theme in dct_competence[name_chapter]:
                            # если тема есть внутри раздела то добавляем ОК и ПК
                            dct_competence[name_chapter][theme].append(competence)
                        else:
                            dct_competence[name_chapter] = {theme:[competence]}



def extract_data_part_themes(df:pd.DataFrame):
    """
    Функция для создания словаря содержащего в себе данные в каких разделах и темах используется та или иная ОК или ПК
    :param df: датафрейм с данными с листа Планируемые результаты
    :return: словарь формата {Наименование раздела (или Наименование раздела темы): используемые там ОК и ПК}
    """
    dct_competence = {} # словарь для хранения данных
    for row in df.itertuples():
        if row[1] not in ('ОК','ПК'): # отсекаем обозначения разделов
            find_part_themes(row[4],dct_competence,row[1]) # находим в каких разделах и темах используется данная ОК или ПК

    return dct_competence






def create_RP_for_UD_OOD(template_work_program:str,data_work_program:str,end_folder:str):
    """
    Скрипт для генерации рабочей програамы для учебной дисциплины ООД (общеобразовательной) с помощью DocxTemplate
    :param template_work_program: путь к шаблону рабочей программы в формате docx
    :type template_work_program: str
    :param data_work_program: путь к файлу с данными для рабочей программы в формате xlsx
    :type data_work_program: str
    :param end_folder: путь к папке куда будет сохранен созданный файл рабочей программы
    :type end_folder: str
    """
    # названия листов
    desc_rp = 'Описание РП'
    structure = 'Объем УД'
    plan = 'План УД'
    content = 'Содержание'
    target = 'Цели'
    result = 'Результаты'
    uupd = 'УУПД'
    main_publ = 'ОИ'
    second_publ = 'ДИ'
    ii_publ = 'ИИ'
    control = 'Контроль'
    plan_results = 'Планируемые результаты'
    # необходимые колонки
    try:
        etalon_cols_lst = [desc_rp,plan,structure,target,result,uupd,content,main_publ,second_publ,ii_publ,control,plan_results]
        etalon_cols = set(etalon_cols_lst)
        temp_wb = openpyxl.load_workbook(data_work_program,read_only=True)
        file_cols = set(temp_wb.sheetnames)
        diff_cols = etalon_cols - file_cols
        temp_wb.close()
        if len(diff_cols) != 0:
            raise DiffSheet

        # Обрабатываем лист Описание РП
        df_desc_rp = pd.read_excel(data_work_program, sheet_name=desc_rp, nrows=1, usecols='A:L')  # загружаем датафрейм
        df_desc_rp.fillna('НЕ ЗАПОЛНЕНО !!!', inplace=True)  # заполняем не заполненные разделы
        df_desc_rp.columns = ['Тип_программы', 'Название_дисциплины', 'Цикл', 'Перечень', 'Код_специальность_профессия',
                              'Год', 'Разработчик', 'Методист', 'ПЦК', 'Пред_ПЦК', 'Должность', 'Утверждающий']

        # Обрабатываем лист Объем УД
        # Открываем файл
        wb = openpyxl.load_workbook(data_work_program, read_only=True)
        target_value = 'итог'

        # Поиск значения в выбранном столбце
        column_number = 1  # Номер столбца, в котором ищем значение (например, столбец A)
        target_row = None  # Номер строки с искомым значением

        for row in wb['Объем УД'].iter_rows(min_row=1, min_col=column_number, max_col=column_number):
            cell_value = row[0].value
            if target_value in str(cell_value).lower():
                target_row = row[0].row
                break

        if not target_row:
            # если не находим строку в которой есть слово итог то выдаем исключение
            print('Не найдена строка с названием Итоговая аттестация в форме ...')

        wb.close()  # закрываем файл

        # если значение найдено то считываем нужное количество строк и  7 колонок
        df_structure = pd.read_excel(data_work_program, sheet_name=structure, nrows=target_row,
                                     usecols='A:C', dtype=str)

        df_structure.iloc[:-1, 1:3] = df_structure.iloc[:-1, 1:3].applymap(convert_to_int)
        df_structure.columns = ['Вид', 'Всего', 'Практика']
        df_structure.fillna('', inplace=True)
        max_load = df_structure.loc[0, 'Всего']  # максимальная учебная нагрузка
        mand_load = df_structure.loc[1, 'Всего']  # обязательная нагрузка
        prof_load = df_structure.loc[1, 'Практика']  # практическая нагрузка

        sam_df = df_structure[
            df_structure['Вид'] == 'Самостоятельная работа обучающегося (всего)']  # получаем часы самостоятельной работы
        if len(sam_df) == 0:
            print('Проверьте наличие строки Самостоятельная работа обучающегося (всего)')
        sam_load = sam_df.iloc[0, 1]

        """
        Обрабатываем лист Планируемые результаты
        для того чтобы при обработке листа План УД добавить колонку с ОК и ПК
        """
        df_plan_result = pd.read_excel(data_work_program, sheet_name=plan_results, usecols='A:E')
        df_plan_result.dropna(inplace=True, how='all')  # удаляем пустые строки
        df_plan_result.columns = ['ОК_ПК','Общие_рез','Дис_рез','Раздел_тема','Тип']
        df_plan_result.fillna('', inplace=True)
        # создаем словарь где указываем в каких разделах и темах используются данные ОК и ПК
        dct_part_themes = extract_data_part_themes(df_plan_result)


        """
            Обрабатываем лист План УД
            """
        main_df, temp_all_result, temp_part_result = extract_data_plan_ood(data_work_program,
                                                                           plan,dct_part_themes)  # извлекаем данные из датафрейма
        #
        check_error_df = create_check_error_df(temp_part_result)

        """
            Обрабатываем лист Содержание
            """
        df_content = pd.read_excel(data_work_program, sheet_name=content, usecols='A',nrows=1,dtype=str)
        df_content = df_content.applymap(str.strip) # очищаем от пробелов

        """
            Обрабатываем листы Цели,Результаты, УУПД
            """
        #Цели
        df_target = pd.read_excel(data_work_program, sheet_name=target, usecols='A')
        lst_target = df_target['Цели'].dropna().tolist()
        lst_target = processing_punctuation_end_string(lst_target, ';\n', '- ', '.')
        # Результаты
        # Личностные результаты
        df_result = pd.read_excel(data_work_program, sheet_name=result, usecols='A:C')
        lst_perconal_result = df_result['Личностные_результаты'].dropna().tolist()
        lst_perconal_result = processing_punctuation_end_string(lst_perconal_result, ';\n', '- ', '.')
        # Метапредметные результаты
        lst_meta_result = df_result['Метапредметные_результаты'].dropna().tolist()
        lst_meta_result = processing_punctuation_end_string(lst_meta_result, ';\n', '- ', '.')
        # Предметные результаты
        lst_predmet_result = df_result['Предметные_результаты'].dropna().tolist()
        lst_predmet_result = processing_punctuation_end_string(lst_predmet_result, ';\n', '- ', '.')
        # УУПД
        df_uupd = pd.read_excel(data_work_program, sheet_name=uupd, usecols='A')
        lst_uupd = df_uupd.iloc[:,0].dropna().tolist()
        lst_uupd = processing_punctuation_end_string(lst_uupd, ';\n', '- ', '.')

        """
            Обрабатываем лист Основные источники
            """

        df_main_publ = pd.read_excel(data_work_program, sheet_name=main_publ, usecols='A:G')
        if df_main_publ.shape[0] != 0:
            df_main_publ.dropna(inplace=True, thresh=1)  # удаляем пустые строки
            df_main_publ.fillna('Не заполнено !!!', inplace=True)
            df_main_publ = df_main_publ.applymap(str)  # приводим к строковому виду
            df_main_publ = df_main_publ.applymap(str.strip)  # очищаем от пробелов в начале и конце

            df_main_publ['Основной_источник'] = df_main_publ.apply(processing_publ, axis=1)  # формируем строку
            df_main_publ.sort_values(by='Основной_источник', inplace=True)
            lst_main_source = df_main_publ['Основной_источник'].tolist()
        else:
            lst_main_source = 'Не заполнено'

        """
            Обрабатываем лист дополнительные источники
            """
        df_second_publ = pd.read_excel(data_work_program, sheet_name=second_publ, usecols='A:G')
        if df_second_publ.shape[0] != 0:
            df_second_publ.dropna(inplace=True, thresh=1)  # удаляем пустые строки
            df_second_publ.fillna('Не заполнено !!!', inplace=True)
            df_second_publ = df_second_publ.applymap(str)  # приводим к строковому виду
            df_second_publ = df_second_publ.applymap(str.strip)  # очищаем от пробелов в начале и конце
            df_second_publ['Основной_источник'] = df_second_publ.apply(processing_publ, axis=1)  # формируем строку
            df_second_publ.sort_values(by='Основной_источник', inplace=True)
            lst_slave_source = df_second_publ['Основной_источник'].tolist()
        else:
            lst_slave_source = 'Не заполнено'

        """
            Обрабатываем лист интернет источники
            """
        df_ii_publ = pd.read_excel(data_work_program, sheet_name=ii_publ, usecols='A:B')
        if df_ii_publ.shape[0]:
            df_ii_publ.dropna(inplace=True, thresh=1)  # удаляем пустые строки

            df_ii_publ.sort_values(by='Название', inplace=True)

            lst_inet_source = insert_type_source(df_ii_publ.copy())
        else:
            lst_inet_source = 'Не заполнено'

        """
            Обрабатываем лист Контроль и Оценка
            """
        df_control = pd.read_excel(data_work_program, sheet_name=control, usecols='A')
        df_control.dropna(inplace=True, thresh=1)  # удаляем пустые строки
        df_control.columns = ['Результаты_обучения']
        _lst_result_educ = df_control['Результаты_обучения'].dropna().tolist()  # создаем список
        if 'Знания:' not in _lst_result_educ:
            messagebox.showerror('Диана Создание рабочих программ','На листе Контроль в первой колонке должно быть слово Знание:\n'
                                                                   'Посмотрите пример в исходном шаблоне')
        border_divide = _lst_result_educ.index('Знания:')  # граница разделения
        lst_skill = _lst_result_educ[1:border_divide]  # получаем список умений
        lst_knowledge = _lst_result_educ[border_divide + 1:]  # получаем список знаний

        lst_skill = processing_punctuation_end_string(lst_skill, ';\n', '- ', '.')  # форматируем выходную строку
        lst_knowledge = processing_punctuation_end_string(lst_knowledge, ';\n', '- ', '.')  # форматируем выходную строку
        df_control.fillna('', inplace=True)







        # Конвертируем датафрейм с описанием программы в список словарей и добавляем туда нужные элементы
        data_program = df_desc_rp.to_dict('records')
        context = data_program[0]
        context['План_УД'] = main_df.to_dict('records')  # содержание учебной дисциплины
        context['Учебная_работа'] = df_structure.to_dict('records')
        context['Контроль_оценка'] = df_control.to_dict('records')
        context['Знать'] = lst_knowledge
        context['Уметь'] = lst_skill

        context['План_результаты'] = df_plan_result.to_dict('records')

        # добавляем единичные переменные
        context['Макс_нагрузка'] = max_load
        context['Обяз_нагрузка'] = mand_load
        context['Проф_направленность'] = prof_load
        context['Сам_работа'] = sam_load

        # Листы Цели,результаты,УУПД
        context['Цели'] = lst_target
        context['Личностные_результаты'] =lst_perconal_result
        context['Метапредметные_результаты'] =lst_meta_result
        context['Предметные_результаты'] =lst_predmet_result
        context['УУПД'] =lst_uupd

        # Лист Содержание
        context['Содержание_дисциплины'] = df_content.iloc[0,0]
        # лист Учебные издания
        context['Основные_источники'] = lst_main_source
        context['Дополнительные_источники'] = lst_slave_source
        context['Интернет_источники'] = lst_inet_source

        doc = DocxTemplate(template_work_program)
        # Получаем ключи используемые в шаблоне
        set_of_variables = doc.get_undeclared_template_variables()

        # Создаем документ
        doc.render(context)
        # сохраняем документ
        # название программы
        name_rp = df_desc_rp['Название_дисциплины'].tolist()[0]
        name_rp = re.sub(r'[\r\b\n\t<> :"?*|\\/]', '_', name_rp)  # очищаем от некорректных символов
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        doc.save(f'{end_folder}/РП ООД {name_rp[:40]} {current_time}.docx')

        # Сохраняем таблицу с результатом проверки часов
        dct_write = {'Sheet1':check_error_df}
        write_index = False
        wb = write_df_to_excel(dct_write, write_index)
        wb.save(f'{end_folder}/Проверка часов УД {name_rp[:40]} {current_time}.xlsx')
    except DiffSheet:
        messagebox.showerror('Диана Создание рабочих программ',
                             f'В таблице не найдены листы {diff_cols},\n'
                             f'В таблице должны быть листы {etalon_cols_lst}\n'
                             f'Возможно вы используете шаблон РП для УД а не шаблон РП для ООД')
    except NotDataMdk as e:
        messagebox.showerror('Диана Создание рабочих программ',
                             f'Лист План УД не заполнен !!!')
    except ControlSemestr as e:
        messagebox.showerror('Диана Создание рабочих программ',
                             f'При обработке листа с Планом УД не найдено слово семестр в первой колонке Курс/семестр\n'
                             f'Должны быть указаны семестры в формате: 2 курс 3 семестр')
    except ControlChapter as e:
        messagebox.showerror('Диана Создание рабочих программ',
                             f'При обработке листа с Планом УД не найдено слово Раздел во второй колонке Раздел\n'
                             f'Должны быть указаны разделы в формате: Раздел 1. Наименование раздела и т.п.')
    except KeyError as e:
        messagebox.showerror('Диана Создание рабочих программ',
                             f'В таблице не найдена колонка с названием {e.args}!\nПроверьте написание названия колонки')
    except ValueError as e:
        messagebox.showerror('Диана Создание рабочих программ',
                             f'В таблице не найден лист,колонка или значение {e.args}!\nПроверьте написание названий')

    except FileNotFoundError:
        messagebox.showerror('Диана Создание рабочих программ',
                             f'Перенесите файлы которые вы хотите обработать в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам')

    except exceptions.TemplateSyntaxError:
        messagebox.showerror('Диана Создание рабочих программ',
                             f'Ошибка в оформлении вставляемых значений в шаблоне\n'
                             f'Проверьте свой шаблон на наличие следующих ошибок:\n'
                             f'1) Вставляемые значения должны быть оформлены двойными фигурными скобками\n'
                             f'{{{{Вставляемое_значение}}}}\n'
                             f'2) В названии колонки в таблице откуда берутся данные - есть пробелы,цифры,знаки пунктуации и т.п.\n'
                             f'в названии колонки должны быть только буквы и нижнее подчеркивание.\n'
                             f'{{{{Дата_рождения}}}}')
    else:
        messagebox.showinfo('Диана Создание рабочих программ', 'Данные успешно обработаны')

if __name__ == '__main__':

    template_work_program = 'data/Шаблон автозаполнения ООД 19_10_24.docx'
    # data_work_program = 'data/ПРИМЕР заполнения таблицы  ООД 13_09.xlsx'
    data_work_program = 'data/Физическая культура.xlsx'
    end_folder = 'data'


    create_RP_for_UD_OOD(template_work_program, data_work_program, end_folder)
    print('Lindy Booth')


