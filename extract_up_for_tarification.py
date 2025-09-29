"""
Скрипт для извлечения данных о дисциплинах находящихся в учебном плане скачанного с сетевого города для использования при подсчете тарификации
"""
from support_function_for_diana import write_df_to_excel,del_sheet
import numpy as np
import pandas as pd
pd.options.display.width= None
pd.options.display.max_columns= None
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import time
import re
import copy
import os
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action="ignore", category=pd.errors.PerformanceWarning)
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None

from tkinter import messagebox


class NotCorrectParams(Exception):
    """
    Исключение для случаев когда нет ни одного корректного параметра
    """
    pass

class NotFile(Exception):
    """
    Обработка случаев когда нет файлов в папке
    """
    pass


def save_convert_to_int(value):
    if isinstance(value,str):
        if value.isdigit():
            return int(value)
        else:
            return value
    else:
        return value


def load_file_params(file_params:str):
    """
    Функция для обработки файла с параметрами
    :param file_params:файл с параметрами
    :return:словарь с параметрами
    """
    try:
        params_df = pd.read_excel(file_params, usecols='A:B')
    except:
        messagebox.showerror('Диана',
                             'Не удалось обработать файл с параметрами обработки учебных планов!\n'
                             'Проверьте файл на повреждения. Пересохраните в новом файле.')
    params_df.dropna(how='any', inplace=True)  # очищаем от неполных строк
    params_df.columns = ['Параметр', 'Значение']  # переименовываем
    # Приводим к нужным типам данных
    # params_df[['Название листа', 'Диапазон']] = params_df[['Название листа', 'Диапазон']].astype(str)
    # params_df['Количество колонок'] = params_df['Количество колонок'].apply(convert_to_int)


def processing_data_up_for_tarification(data_folder:str, file_params:str, result_folder:str):
    """
    Функция для извлечения данных по учебным планам
    :param data_folder: папка с данными
    :param file_params: файл с параметрами
    :param result_folder: конечная папка
    """
    error_df = pd.DataFrame(
        columns=['Название файла','Описание ошибки'])  # датафрейм для ошибок

    dct_params = load_file_params(file_params) # получаем параметры # TODO Доделать функцию
    name_sheet = 'Учебный план' # название листа
    quantity_header = 6 # количество строк заголовка
    number_main_column = 2 - 1 # порядковый номер колонки с наименованиями
    quantity_cols = 15 # количество колонок с данными которые нужно собрать

    finish_df = pd.DataFrame(columns=list(range(quantity_cols+1))) # создаем итоговый датафрейм куда будут добавляться все данные

    for dirpath, dirnames, filenames in os.walk(data_folder):
        for file in filenames:
            if file.endswith('.xls') or file.endswith('.ods'):
                temp_error_df = pd.DataFrame(
                    data=[[f'{file}',
                           f'Программа обрабатывает файлы с разрешением xlsx. XLS и ODS файлы не обрабатываются !'
                           ]],
                    columns=['Название файла',
                             'Описание ошибки'])
                error_df = pd.concat([error_df, temp_error_df], axis=0,
                                     ignore_index=True)
                continue
            if not file.startswith('~$') and file.endswith('.xlsx'):
                name_file = file.split('.xlsx')[0]
                print(name_file)  # обрабатываемый файл
                try:
                    temp_df = pd.read_excel(f'{dirpath}/{file}',sheet_name=name_sheet,skiprows=quantity_header,header=None,usecols=list(range(0,quantity_cols)))  # открываем файл
                except:
                    temp_error_df = pd.DataFrame(
                        data=[[f'{file}',
                               f'Не удалось обработать файл. Возможно файл поврежден'
                               ]],
                        columns=['Название файла',
                                 'Описание ошибки'])
                    error_df = pd.concat([error_df, temp_error_df], axis=0,
                                         ignore_index=True)
                    continue

                dct_file = dict()  # основной словарь со структорой Название файла:{Дисциплина:список значений}

                temp_df.dropna(subset=[number_main_column],inplace=True) # убираем пустые строки в колонке с наименованиями
                temp_df.insert(0,'Имя файла',name_file) # добавляем название файла
                temp_df.columns = list(range(quantity_cols+1)) # переименовываем колонки

                finish_df = pd.concat([finish_df,temp_df])



    t = time.localtime()  # получаем текущее время
    current_time = time.strftime('%H_%M_%S', t)

    # Сохраняем ошибки
    wb = openpyxl.Workbook()
    for r in dataframe_to_rows(error_df, index=False, header=True):
        wb['Sheet'].append(r)

    wb['Sheet'].column_dimensions['A'].width = 30
    wb['Sheet'].column_dimensions['B'].width = 40
    wb['Sheet'].column_dimensions['C'].width = 50

    wb.save(f'{result_folder}/ОШИБКИ от {current_time}.xlsx')

    finish_df.sort_values(by=number_main_column,inplace=True)
    lst_finish_cols = ['Название файла']
    lst_finish_cols.extend(list(range(1,quantity_cols+1)))
    finish_df.columns = lst_finish_cols # присваиваем более понятный названия колонок

    finish_df = finish_df.applymap(save_convert_to_int)

    temp_wb = write_df_to_excel(
        {'Общий свод': finish_df},
        write_index=False)
    temp_wb = del_sheet(temp_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
    temp_wb.save(f'{result_folder}/Общий результат {current_time}.xlsx')

    # Сохраняем по отдельным файлам
    name_column = finish_df.columns[number_main_column+1]  # получаем название колонки
    lst_value_column = finish_df.iloc[:,number_main_column+1].unique()
    lst_value_column = [value for value in lst_value_column if pd.notna(value)]
    lst_value_column = list(map(str, lst_value_column))

    used_name_file = set()  # множество для уже использованных имен файлов
    for idx, value in enumerate(lst_value_column):
        wb = openpyxl.Workbook()  # создаем файл
        temp_df = finish_df[finish_df[name_column] == value]  # отфильтровываем по значению
        short_name = value[:40]  # получаем обрезанное значение
        short_name = re.sub(r'[\r\b\n\t\'+()<> :"?*|\\/]', '_', short_name)
        if short_name.lower() in used_name_file:
            short_name = f'{short_name}_{idx}'  # добавляем окончание
        for row in dataframe_to_rows(temp_df, index=False, header=True):
            wb['Sheet'].append(row)

        # Устанавливаем автоширину для каждой колонки
        for column in wb['Sheet'].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            if adjusted_width >= 80:
                wb['Sheet'].column_dimensions[column_name].width = 80
                for cell in wb['Sheet'][column_name]:
                    cell.alignment = Alignment(horizontal='left',wrap_text=True)
            else:
                wb['Sheet'].column_dimensions[column_name].width = adjusted_width+3

        finish_path = f'{result_folder}/По отдельности'
        if not os.path.exists(finish_path):
            os.makedirs(finish_path)

        wb.save(f'{finish_path}/{short_name}.xlsx')
        used_name_file.add(short_name.lower())
        wb.close()












if __name__ == '__main__':
    main_data_folder = 'data/Учебные планы'
    main_file_params = 'data/Параметры сбора.xlsx'
    main_result_folder = 'data/Результат'
    processing_data_up_for_tarification(main_data_folder,main_file_params,main_result_folder)
    print('Lindy Booth')