"""
Скрипт для создания отчетности по преподавателям
"""
import datetime

from support_function_for_diana import write_df_to_excel,del_sheet
from create_docs import generate_docs
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from docxtpl import DocxTemplate
import string
import time
import re
import os
from tkinter import messagebox
from jinja2 import exceptions

pd.options.mode.chained_assignment = None  # default='warn'
pd.set_option('display.max_columns', None)  # Отображать все столбцы
import warnings

warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.filterwarnings('ignore', category=FutureWarning, module='openpyxl')
warnings.filterwarnings("ignore", category=DeprecationWarning)


def check_required_sheet_in_file(path_to_checked_file: str, func_required_sheets_columns: dict, func_name_file: str):
    """
    Функция для проверки наличия обязательных листов в файле
    :param path_to_checked_file: путь к проверяемому файлу
    :param func_required_sheets_columns: словарь с данными для проверки формата {Лист:[Список обязательных колонок}
    :param func_name_file:  имя проверяемого файла
    :return: датафрейм с найденными ошибками
    """
    # загружаем файл, чтобы получить названия листов
    check_sheets_wb = openpyxl.load_workbook(path_to_checked_file, read_only=True)
    file_sheets = check_sheets_wb.sheetnames  # получаем названия листов
    check_sheets_wb.close()  # закрываем файл
    # проверяем наличие нужных листов
    diff_sheets = set(func_required_sheets_columns.keys()).difference(set(file_sheets))
    if len(diff_sheets) != 0:
        # Записываем ошибку
        temp_error_df = pd.DataFrame(data=[[f'{func_name_file}', ';'.join(diff_sheets),
                                            'Не найдены указанные обязательные листы']],
                                     columns=['Название файла', 'Название листа',
                                              'Описание ошибки'])
        return temp_error_df


def check_required_columns_in_sheet(path_to_checked_file: str, func_required_sheets_columns: dict, func_name_file: str):
    """
    Функция для проверки наличия обязательных колонок на каждом листе в файле
    :param path_to_checked_file: путь к проверяемому файлу
    :param func_required_sheets_columns: словарь с данными для проверки формата {Лист:[Список обязательных колонок}
    :param func_name_file:  имя проверяемого файла
    :return: датафрейм с найденными ошибками
    """
    # датафрейм для сбора ошибок
    check_error_req_columns_df = pd.DataFrame(columns=['Название файла', 'Название листа', 'Описание ошибки'])
    for name_sheet, lst_req_cols in func_required_sheets_columns.items():
        check_cols_df = pd.read_excel(path_to_checked_file, sheet_name=name_sheet)  # открываем файл
        diff_cols = set(lst_req_cols).difference(set(check_cols_df.columns))  # ищем разницу в колонках
        if len(diff_cols) != 0:
            # Записываем ошибку
            temp_error_df = pd.DataFrame(data=[[f'{func_name_file}', name_sheet,
                                                f'На листе не найдены указанные обязательные колонки: {";".join(diff_cols)}']],
                                         columns=['Название файла', 'Название листа',
                                                  'Описание ошибки'])
            check_error_req_columns_df = pd.concat([check_error_req_columns_df, temp_error_df], axis=0,
                                                   ignore_index=True)

    if len(check_error_req_columns_df) != 0:
        return check_error_req_columns_df


def prepare_sheet_general_inf(path_to_file:str, name_sheet:str, lst_columns:list):
    """
    Функция для извлечения данных из листа Общие сведения
    :param path_to_file: путь к файлу
    :param name_sheet: название листа
    :param lst_columns: список колонок из которых будут извлекаться данные
    :return: подготовленный датафрейм
    """
    func_df = pd.read_excel(path_to_file,sheet_name=name_sheet,usecols=lst_columns)
    func_df.dropna(how='all',inplace=True) # удаляем пустые строки
    # очищаем от пробельных символов в начале и конце каждой ячейки
    func_df = func_df.applymap(lambda x:x.strip() if isinstance(x,str) else x)
    person_df = func_df.copy() # создаем датафрейм для сохранения всех строк
    person_df['ФИО'] = person_df.iloc[0,0]
    person_df['Дата рождения'] = person_df.iloc[0,1]
    person_df['Дата начала работы в ПОО'] = person_df.iloc[0,2]
    person_df['Общий стаж работы'] = person_df.iloc[0,4]
    person_df['Педагогический стаж'] = person_df.iloc[0,5]
    person_df['Категория'] = person_df.iloc[0,9]
    person_df['№ приказа на аттестацию, дата'] = person_df.iloc[0,10]
    person_df['Наличие личного сайта, блога (ссылка)'] = person_df.iloc[0,11]
    dis_lst = func_df['Преподаваемая дисциплина'].tolist() # список дисциплин преподавателя
    # список полученных дипломов об образовании
    educ_lst = func_df['Сведения об образовании (образовательное учреждение, квалификация, год окончания)'].tolist()
    func_df = func_df.iloc[0,:].to_frame().transpose()
    func_df.loc[0,'Преподаваемая дисциплина'] = ';'.join(dis_lst) # превращаем в строку список дисциплин
    func_df.loc[0,'Сведения об образовании (образовательное учреждение, квалификация, год окончания)'] = ';'.join(educ_lst) # превращаем в строку список дисциплин
    return func_df,person_df


def prepare_sheet_standart(path_to_file:str, name_sheet:str, lst_columns:list):
    """
    Функция для извлечения данных из листа Повышение квалификации
    :param path_to_file: путь к файлу
    :param name_sheet: название листа
    :param lst_columns: список колонок из которых будут извлекаться данные
    :return: подготовленный датафрейм
    """
    func_df = pd.read_excel(path_to_file,sheet_name=name_sheet,usecols=lst_columns)
    func_df.dropna(how='all',inplace=True) # удаляем пустые строки
    # очищаем от пробельных символов в начале и конце каждой ячейки
    func_df = func_df.applymap(lambda x:x.strip() if isinstance(x,str) else x)

    return func_df


def extract_last_date(value:str):
    """
    Функция для извлечения последней даты из ячейки
    :param value:значение ячейки
    :return: дата в виде строки или None
    """
    result_lst = re.findall(r'\d{2}.\d{2}.\d{4}',value)
    if result_lst:
        return result_lst[-1]
    else:
        return None

def prepare_date(value):
    """
    Функция для конвертации дат с учетом возможных вариантов вида 01.01.2023-02.03.2023
    :param value: значение
    :return: дата в формате timestamp
    """
    try:
        prep_date = pd.to_datetime(value,dayfirst=True) # конвертируем в дату
        return prep_date
    except:
        value_str = str(value) # конвертируем в строку
        prep_date = extract_last_date(value_str) # ищем последнюю дату
        if prep_date:
            prep_date = datetime.datetime.strptime(prep_date,'%d.%m.%Y')  # конвертируем в дату
            return prep_date
        else:
            return None


def selection_by_date(df:pd.DataFrame,start_date:str,end_date:str,name_date_column,name_file:str,name_sheet:str,
                      union_df:pd.DataFrame,error_df:pd.DataFrame,name_teacher:str):
    """
    Функция для отбора тех строк датафрейма что подходят под выбранные даты
    :param df: датафрейм с данными
    :param start_date: стартовая дата в формате строки 01.01.2024
    :param end_date: конечная дата в формате строки 17.12.2024
    :param name_date_column: список названий колонок с датами
    :param name_file:  название файла
    :param name_sheet: название листа с ошибкой
    :param union_df: датафрейм для консолидации
    :param error_df: общий датафрейм с ошибками
    :param name_teacher: ФИО педагога
    :return: отфильтрованный датафрейм
    """
    if name_sheet != 'Общие сведения':
        # конвертируем даты в формат дат


        df['_Отбор даты'] = df[name_date_column].apply(prepare_date)
        date_error_df = df[df['_Отбор даты'].isnull()] # отбираем строки с ошибками в датах
        if len(date_error_df) != 0:
            number_row_error = list(map(lambda x:str(x+2),date_error_df.index)) # получаем номера строк с ошибками прибавляя 2
            temp_error_df = pd.DataFrame(data=[[f'{name_file}', name_sheet,
                                                f'В колонке {name_date_column} в указанных строках неправильно записаны даты: {";".join(number_row_error)}. Требуемый формат: 21.05.2024'
                                                f' или 05.06.2024-15.08.2024']],
                                         columns=['Название файла', 'Название листа',
                                                  'Описание ошибки'])
            error_df = pd.concat([error_df, temp_error_df], axis=0,
                                                   ignore_index=True)


        df = df[df['_Отбор даты'].notnull()] # отбираем строки с правильной датой

        df = df[df['_Отбор даты'].between(start_date,end_date)] # получаем значения в указанном диапазоне
        df[name_date_column] = df[name_date_column].apply(lambda x: x.strftime('%d.%m.%Y') if isinstance(x, (pd.Timestamp, datetime.datetime)) else x)
        df.drop(columns=['_Отбор даты'], inplace=True) # удаляем служебную колонку
        df.insert(0,'ФИО',name_teacher)



    # Соединяем
    union_df = pd.concat([union_df,df])


    return union_df,df, error_df



def create_report_teacher(template_folder:str,data_folder: str, result_folder: str,start_date:pd.Timestamp,end_date:pd.Timestamp):
    """
    Функция для создания отчетности по преподавателям
    :param template_folder: папка, где лежат шаблоны отчетов
    :param data_folder: папка где хранятся личные дела преподавателей
    :param result_folder: папка в которую будут сохраняться итоговые файлы
    :param start_date: начальная дата, если ничего не указано то 01.01.1900
    :param end_date: конечная дата, если ничего не указано то 01.01.2100
    """
    # обязательные листы
    required_sheets_columns = {'Общие сведения': ['ФИО', 'Дата рождения', 'Дата начала работы в ПОО',
                                                  'Преподаваемая дисциплина', 'Общий стаж работы',
                                                  'Педагогический стаж',
                                                  'Сведения об образовании (образовательное учреждение, квалификация, год окончания)',
                                                  'Квалификация','Год окончания',
                                                  'Категория', '№ приказа на аттестацию, дата',
                                                  'Наличие личного сайта, блога (ссылка)'],
                               'Повышение квалификации': ['Название программы повышения квалификации',
                                                          'Вид повышения квалификации',
                                                          'Место прохождения программы',
                                                          'Дата прохождения программы (с какого по какое число, месяц, год)',
                                                          'Количество академических часов',
                                                          'Наименование подтверждающего документа, его номер и дата выдачи'],
                               'Стажировка': ['Место стажировки', 'Кол-во часов', 'Дата'],
                               'Методические разработки': ['Вид методического издания', 'Название издания',
                                                           'Профессия/специальность ',
                                                           'Дата разработки', 'Кем утверждена'],
                               'Мероприятия, пров. ППС': ['Название мероприятия', 'Дата', 'Уровень мероприятия'],
                               'Личное выступление ППС': ['Дата', 'Название мероприятия', 'Тема', 'Вид мероприятия',
                                                          'Уровень мероприятия',
                                                          'Способ участия', 'Результат участия'],
                               'Публикации': ['Полное название статьи', 'Издание', 'Дата выпуска'],
                               'Открытые уроки': ['Вид занятия','Дисциплина', 'Группа', 'Тема', 'Дата проведения'],
                               'Взаимопосещение': ['ФИО посещенного педагога', 'Дата посещения', 'Группа', 'Тема'],
                               'УИРС': ['ФИО обучающегося', 'Профессия/специальность', 'Группа', 'Вид мероприятия',
                                        'Название мероприятия',
                                        'Тема', 'Способ участия', 'Уровень мероприятия', 'Дата проведения',
                                        'Результат участия'],
                               'Работа по НМР': ['Тема НИРП', 'Проведено ли обобщение опыта', 'Форма обобщения опыта',
                                                 'Место обобщения опыта', 'Дата обобщения опыта'],
                               'Данные для списков': []
                               }
    error_df = pd.DataFrame(
        columns=['Название файла', 'Название листа', 'Описание ошибки'])  # датафрейм для ошибок

    teachers_dct = dict() # словарь в котором будут храниться словари с данными листов для каждого преподавателя

    print(start_date)
    print(type(start_date))
    print(end_date)
    print(type(end_date))



    # Создаем датафреймы для консолидации данных
    general_inf_df = pd.DataFrame(columns=required_sheets_columns['Общие сведения'])
    skills_dev_df = pd.DataFrame(columns=required_sheets_columns['Повышение квалификации'])
    skills_dev_df.insert(0,'ФИО','')
    internship_df = pd.DataFrame(columns=required_sheets_columns['Стажировка'])
    internship_df.insert(0, 'ФИО', '')
    method_dev_df = pd.DataFrame(columns=required_sheets_columns['Методические разработки'])
    method_dev_df.insert(0, 'ФИО', '')
    events_teacher_df = pd.DataFrame(columns=required_sheets_columns['Мероприятия, пров. ППС'])
    events_teacher_df.insert(0, 'ФИО', '')
    personal_perf_df = pd.DataFrame(columns=required_sheets_columns['Личное выступление ППС'])
    personal_perf_df.insert(0, 'ФИО', '')
    publications_df = pd.DataFrame(columns=required_sheets_columns['Публикации'])
    publications_df.insert(0, 'ФИО', '')
    open_lessons_df = pd.DataFrame(columns=required_sheets_columns['Открытые уроки'])
    open_lessons_df.insert(0, 'ФИО', '')
    mutual_visits_df = pd.DataFrame(columns=required_sheets_columns['Взаимопосещение'])
    mutual_visits_df.insert(0, 'ФИО', '')
    student_perf_df = pd.DataFrame(columns=required_sheets_columns['УИРС'])
    student_perf_df.insert(0, 'ФИО', '')
    nmr_df = pd.DataFrame(columns=required_sheets_columns['Работа по НМР'])
    nmr_df.insert(0, 'ФИО', '')

    for idx, file in enumerate(os.listdir(data_folder)):
        if not file.startswith('~$') and file.endswith('.xlsx'):
            name_file = file.split('.xlsx')[0]  # Получаем название файла
            path_to_file = f'{data_folder}/{file}'
            # Проверяем наличие нужных листов в файле
            error_req_sheet_df = check_required_sheet_in_file(path_to_file, required_sheets_columns, name_file)
            if error_req_sheet_df is not None:
                error_df = pd.concat([error_df, error_req_sheet_df], axis=0, ignore_index=True)
                continue
            # Проверка наличия нужных колонок в листах
            error_req_columns_sheet_df = check_required_columns_in_sheet(path_to_file, required_sheets_columns,
                                                                         name_file)
            if error_req_columns_sheet_df is not None:
                error_df = pd.concat([error_df, error_req_columns_sheet_df], axis=0, ignore_index=True)
                continue
            print(name_file)
            # Обрабатываем лист Общие сведения
            teachers_dct[name_file] = {key: pd.DataFrame() for key in required_sheets_columns.keys()} # Создаем ключи для конкретного преподавателя
            prep_general_inf_df,teachers_dct[name_file]['Общие сведения'] = prepare_sheet_general_inf(path_to_file, 'Общие сведения', required_sheets_columns['Общие сведения'])
            fio_teacher = teachers_dct[name_file]['Общие сведения'].iloc[0,0] # получаем ФИО преподавателя для добавления в
            # Отбираем данные по датам
            general_inf_df, _, error_df = selection_by_date(prep_general_inf_df,start_date,end_date,'Дата начала работы в ПОО',
                                                        name_file,'Общие сведения',general_inf_df,error_df,fio_teacher)
            # Обрабатываем лист Повышение квалификации
            prep_skills_dev_df = prepare_sheet_standart(path_to_file, 'Повышение квалификации', required_sheets_columns['Повышение квалификации'])
            # сохраняем в датафрейм
            skills_dev_df,teachers_dct[name_file]['Повышение квалификации'],error_df = selection_by_date(prep_skills_dev_df,start_date,end_date,'Дата прохождения программы (с какого по какое число, месяц, год)',
                                                        name_file,'Повышение квалификации',skills_dev_df,error_df,fio_teacher)
            # Обрабатываем лист Стажировка
            prep_internship_df = prepare_sheet_standart(path_to_file, 'Стажировка', required_sheets_columns['Стажировка'])
            internship_df,teachers_dct[name_file]['Стажировка'],error_df = selection_by_date(prep_internship_df,start_date,end_date,'Дата',
                                                        name_file,'Стажировка',internship_df,error_df,fio_teacher)

            # Обрабатываем лист Методические разработки
            prep_method_dev_df = prepare_sheet_standart(path_to_file, 'Методические разработки', required_sheets_columns['Методические разработки'])
            method_dev_df,teachers_dct[name_file]['Методические разработки'],error_df = selection_by_date(prep_method_dev_df,start_date,end_date,'Дата разработки',
                                                        name_file,'Методические разработки',method_dev_df,error_df,fio_teacher)

            # Обрабатываем лист Мероприятия, пров. ППС
            prep_events_teacher_df = prepare_sheet_standart(path_to_file, 'Мероприятия, пров. ППС', required_sheets_columns['Мероприятия, пров. ППС'])
            events_teacher_df,teachers_dct[name_file]['Мероприятия, пров. ППС'], error_df = selection_by_date(prep_events_teacher_df, start_date, end_date, 'Дата',
                                                        name_file, 'Мероприятия, пров. ППС', events_teacher_df, error_df,fio_teacher)

            # Обрабатываем лист Личное выступление ППС
            prep_personal_perf_df = prepare_sheet_standart(path_to_file, 'Личное выступление ППС', required_sheets_columns['Личное выступление ППС'])
            personal_perf_df,teachers_dct[name_file]['Личное выступление ППС'], error_df = selection_by_date(prep_personal_perf_df, start_date, end_date, 'Дата',
                                                            name_file, 'Личное выступление ППС', personal_perf_df,
                                                            error_df,fio_teacher)

            # Обрабатываем лист Публикации
            prep_publications_df = prepare_sheet_standart(path_to_file, 'Публикации', required_sheets_columns['Публикации'])
            publications_df,teachers_dct[name_file]['Публикации'], error_df = selection_by_date(prep_publications_df, start_date, end_date, 'Дата выпуска',
                                                           name_file, 'Публикации', publications_df,
                                                           error_df,fio_teacher)
            # Обрабатываем лис Открытые уроки
            prep_open_lessons_df = prepare_sheet_standart(path_to_file, 'Открытые уроки', required_sheets_columns['Открытые уроки'])
            open_lessons_df, teachers_dct[name_file]['Открытые уроки'],error_df = selection_by_date(prep_open_lessons_df, start_date, end_date, 'Дата проведения',
                                                          name_file, 'Открытые уроки', open_lessons_df,
                                                          error_df,fio_teacher)
            # Обрабатываем лист Взаимопосещение
            prep_mutual_visits_df = prepare_sheet_standart(path_to_file, 'Взаимопосещение', required_sheets_columns['Взаимопосещение'])
            mutual_visits_df,teachers_dct[name_file]['Взаимопосещение'] ,error_df = selection_by_date(prep_mutual_visits_df, start_date, end_date, 'Дата посещения',
                                                          name_file, 'Взаимопосещение', mutual_visits_df,
                                                          error_df,fio_teacher)
            # Обрабатываем лист УИРС
            prep_student_perf_df = prepare_sheet_standart(path_to_file, 'УИРС', required_sheets_columns['УИРС'])
            student_perf_df,teachers_dct[name_file]['УИРС'] ,error_df = selection_by_date(prep_student_perf_df, start_date, end_date,
                                                           'Дата проведения',
                                                           name_file, 'УИРС', student_perf_df,
                                                           error_df,fio_teacher)
            # Обрабатываем лист Работа по НМР
            prep_nmr_df = prepare_sheet_standart(path_to_file, 'Работа по НМР', required_sheets_columns['Работа по НМР'])
            nmr_df,teachers_dct[name_file]['Работа по НМР'] ,error_df = selection_by_date(prep_nmr_df, start_date, end_date,
                                                          'Дата обобщения опыта',
                                                          name_file, 'Работа по НМР', nmr_df,
                                                          error_df,fio_teacher)




    # генерируем текущее время
    t = time.localtime()
    current_time = time.strftime('%H_%M_%S', t)
    # Сохраняем файл эксель с данными
    # Словарь для передачи в функцию форматирования

    dct_df = {'Общие сведения':general_inf_df,'Повышение квалификации':skills_dev_df,
              'Стажировка':internship_df,'Методические разработки':method_dev_df,'Мероприятия, пров. ППС':events_teacher_df,
              'Личное выступление ППС':personal_perf_df,'Публикации':publications_df,'Открытые уроки':open_lessons_df,
              'Взаимопосещение':mutual_visits_df,'УИРС':student_perf_df,'Работа по НМР':nmr_df}
    main_wb = write_df_to_excel(dct_df,write_index=False)
    del_sheet(main_wb,['Sheet'])
    main_wb.save(f'{result_folder}/Свод {current_time}.xlsx')


    # Сохраняем файл с ошибками
    error_wb = write_df_to_excel({'Ошибки':error_df},write_index=False)
    del_sheet(error_wb,['Sheet'])
    error_wb.save(f'{result_folder}/Ошибки {current_time}.xlsx')

    # Сохраняем файлы в формате docx
    generate_docs({'Личные дела':teachers_dct,'Отчет':dct_df},template_folder,result_folder) # создаем личные дела




if __name__ == '__main__':
    main_data_folder = 'data/Данные'
    main_result_folder = 'data/Результат'
    main_template_folder = 'data/Шаблоны'
    main_start_date = '06.06.1900'
    main_end_date = '01.05.2100'

    create_report_teacher(main_template_folder ,main_data_folder,main_result_folder, main_start_date, main_end_date)

    print('Lindy Booth')
