"""
скрипт для создания рабочих программ производственных практик по профмодулям
"""
import pandas as pd
import numpy as np
import openpyxl
from docxtpl import DocxTemplate
import string
import time
import re
from tkinter import messagebox
from jinja2 import exceptions

pd.options.mode.chained_assignment = None  # default='warn'
pd.set_option('display.max_columns', None)  # Отображать все столбцы
pd.set_option('display.expand_frame_repr', False)  # Не переносить строки
import warnings

warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.filterwarnings('ignore', category=FutureWarning, module='openpyxl')
warnings.filterwarnings("ignore", category=DeprecationWarning)

class ControlWord_PM(Exception):
    """
    Исключение для случая когда на листе Контроль шаблона ПМ пропущены слова Умения: ,Знания:, Практический опыт:
    """
    pass

class DiffSheet(Exception):
    """
    Исключение для случаев когда отсутствуют нужные колонки в файле
    """
    pass

def convert_to_int(cell):
    """
    Метод для проверки значения ячейки
    :param cell: значение ячейки
    :return: число в формате int
    """
    if cell is np.nan:
        return ''
    try:
        value = float(cell)
        return int(value)
    except:
        return cell


def extract_lr(cell):
    """
    Функция для создания 2 списков из данных одной колонки
    первая колонка это код личностного результата а вторая колонка  это описание
    извлечение будет с помощью регулярок
    :param df: датафрем из одной колонки
    :return:
    """
    value = str(cell)  # делаем строковй
    result = re.split(r'(\d+\.\s*?)', value)
    result = [value for value in result if value]  # убираем пустые значения
    if len(result) >= 3:
        end_lr = result[1].strip()
        end_lr = end_lr.rstrip(string.punctuation)  # очищаем от точки в конце
        return f'{result[0]}{end_lr}'
    else:
        lr = result[0].strip()
        lr = lr.rstrip(string.punctuation)
        return lr


def extract_descr_lr(cell):
    """
    Функция для создания 2 списков из данных одной колонки
    первая колонка это код личностного результата а вторая колонка  это описание
    извлечение будет с помощью регулярок
    :param df: датафрем из одной колонки
    :return:
    """
    value = str(cell)  # делаем строковй
    result = re.split(r'(\d+\.\s*?)', value)
    result = [value for value in result if value]
    if len(result) >= 3:
        descr_lr = result[2].strip()
        descr_lr = descr_lr.rstrip(string.punctuation)  # очищаем от точки в конце
        return f'{descr_lr}.'
    else:
        return ''


def processing_punctuation_end_string(lst_phrase: list, sep_string: str, sep_begin: str, sep_end: str) -> str:
    """
    Очистка каждого элемента списка от знаков пунктации в конце,
    добавление разделитея, добавление точки в конце
    :param lst_phrase: список элементов
    :param sep_string: разделитель между элементами списка
    :param sep_begin: начальный разделитель
    :param sep_end: знак пунктуации в конце
    :return: строку с разделителями и переносами строки
    """
    if len(lst_phrase) == 0:
        return ''
    lst_phrase = list(map(str, lst_phrase))
    temp_lst = list(map(lambda x: sep_begin + x, lst_phrase))
    temp_lst = list(map(lambda x: x.strip(), temp_lst))  # очищаем от прбельных символов в начале и конце
    temp_lst = list(map(lambda x: x.rstrip(string.punctuation), temp_lst))  # очищаем от знаков пунктуации
    temp_lst[-1] = temp_lst[-1] + sep_end  # добавляем конечный знак пунктуации
    temp_str = f'{sep_string}'.join(temp_lst)  # создаем строку с разделителями
    return temp_str


def insert_type_source(df: pd.DataFrame) -> list:
    """
    Вставка в строку слов [Электронный ресурс] Форма доступа:
    :param lst_phrase:датафрейм
    :return: список измененных строк
    """
    out_lst = []  # список для хранения строк
    for row in df.itertuples():
        name = str(row[1])
        url_ii = str(row[2])
        name = name.strip()  # очищаем от пробельных символов
        name = name.rstrip(string.punctuation)  # очищаем от знаков препинания
        name = name.strip()
        url_ii = url_ii.strip()  # очищаем от пробельных символов
        url_ii = url_ii.rstrip(string.punctuation)  # очищаем от знаков препинания
        url_ii = url_ii.strip()
        temp_str = f'{name} [Электронный ресурс] Форма доступа:{url_ii}'
        out_lst.append(temp_str)

    return out_lst


def processing_publ(row):
    """
    Функция для генерации строки с литературой в нужном формате
    :param row: строка датафрейма
    :return: сумма строк датафрейма в нужном формате
    """
    author = row[0]  # автор(ы)
    name_book = row[1]  # название
    full_city = row[2]  # полное название города
    short_city = row[3]  # краткое название города
    publ_house = row[4]  # издательство
    year = row[5]  # год издания
    quan_pages = row[6]  # число страниц
    author = author.rstrip(string.punctuation)  # очищаем от символа пунктуации в конце
    name_book = name_book.rstrip(string.punctuation)  # очищаем от символа пунктуации в конце
    short_city = short_city.rstrip(string.punctuation)  # очищаем от символа пунктуации в конце
    publ_house = publ_house.rstrip(string.punctuation)  # очищаем от символа пунктуации в конце
    # извлекаем год
    result = re.search(r'\d{4}', year)
    if result:
        clean_year = result.group()
    else:
        clean_year = 'Неправильно заполнен год издания, введите год в формате 4 цифры без букв'

    # извлекаем количество страниц
    result = re.search(r'\d+', quan_pages)
    if result:
        clean_quan_pages = result.group()
    else:
        clean_quan_pages = 'Неправильно заполнено количество страниц, введите количество в виде числа без букв'
    # Формируем итоговую строку
    out_str = f'{author}. {name_book}.- {short_city}.: {publ_house}, {clean_year}.- {clean_quan_pages} c.'

    return out_str


def sum_column_any_value(df:pd.DataFrame,name_column:str):
    """
    Суммирование колонки с разными типами значений в том числе строковыми.
    """
    lst_value = df[name_column].dropna().tolist()
    sum_value = [value for value in lst_value if isinstance(value,(int,float))] # отбираем только числа
    return sum(sum_value) # возвращаем сумму

def extract_data_mdk(data_pm,sheet_name):
    """
    Функция для получения датафрейма из листа файла
    :param data_pm: путь к файлу
    :param sheet_name: имя листа
    :return: датафрейм
    """
    lst_type_lesson = ['урок', 'практическое занятие', 'лабораторное занятие',
                       'курсовая работа (КП)']  # список типов занятий
    dct_all_sum_result = {key: 0 for key in lst_type_lesson}  # создаем словарь для подсчета значений



    df_plan_pm = pd.read_excel(data_pm,sheet_name=sheet_name,skiprows=1, usecols='A:H')
    df_plan_pm.dropna(inplace=True, thresh=1)  # удаляем пустые строки

    df_plan_pm.columns = ['Курс_семестр', 'Раздел', 'Тема', 'Содержание', 'Количество_часов', 'Практика', 'Вид_занятия',
                          'СРС']
    df_plan_pm['Курс_семестр'].fillna('Пусто', inplace=True)
    df_plan_pm['Раздел'].fillna('Пусто', inplace=True)
    df_plan_pm['Тема'].fillna('Пусто', inplace=True)

    # Считаем общие суммы
    mdk_all_sum = int(sum_column_any_value(df_plan_pm, 'Количество_часов'))  # получаем сумму общие часы
    mdk_all_prac_sum = int(sum_column_any_value(df_plan_pm, 'Практика'))  # получаем сумму общие часы
    mdk_all_srs_sum = int(sum_column_any_value(df_plan_pm, 'СРС'))  # сумма срс
    for type_lesson in lst_type_lesson:
        _df = df_plan_pm[df_plan_pm['Вид_занятия'] == type_lesson]  # фильтруем датафрейм
        dct_all_sum_result[type_lesson] = int(sum_column_any_value(_df, 'Количество_часов'))  # получаем значение

    dct_all_sum_result['Всего часов'] = mdk_all_sum
    dct_all_sum_result['Всего практики'] = mdk_all_prac_sum
    dct_all_sum_result['Всего СРС'] = mdk_all_srs_sum


    borders = df_plan_pm[
        df_plan_pm['Курс_семестр'].str.contains('семестр')].index  # получаем индексы строк где есть слово семестр

    part_df = []  # список для хранения кусков датафрейма
    previos_border = -1
    # делим датафрем по границам
    for value_border in borders:
        part = df_plan_pm.iloc[previos_border:value_border]
        part_df.append(part)
        previos_border = value_border

    # добавляем последнюю часть
    last_part = df_plan_pm.iloc[borders[-1]:]
    part_df.append(last_part)

    part_df.pop(0)  # удаляем нулевой элемент так как он пустой

    main_df = pd.DataFrame(
        columns=['Курс_семестр', 'Раздел', 'Тема', 'Содержание', 'Количество_часов', 'Практика', 'Вид_занятия',
                 'СРС'])  # создаем базовый датафрейм

    for df in part_df:
        dct_sum_result = {key: 0 for key in lst_type_lesson}  # создаем словарь для подсчета значений
        for type_lesson in lst_type_lesson:
            _df = df[df['Вид_занятия'] == type_lesson]  # фильтруем датафрейм
            _df['Количество_часов'].fillna(0, inplace=True)
            _df['Количество_часов'] = _df['Количество_часов'].astype(int)
            dct_sum_result[type_lesson] = _df['Количество_часов'].sum()
        # создаем строку с описанием
        margint_text = 'Итого часов за семестр:\nиз них\nтеория\nпрактические занятия\nлабораторные занятия\nкурсовая работа (КП)'

        all_hours = sum(dct_sum_result.values())  # общая сумма часов

        theory_hours = dct_sum_result['урок']  # часы теории
        praktice_hours = dct_sum_result['практическое занятие']  # часы практики
        lab_hours = dct_sum_result['лабораторное занятие']  # часы лабораторных
        kurs_hours = dct_sum_result['курсовая работа (КП)']  # часы курсовых

        value_text = f'{all_hours}\n \n{theory_hours}\n{praktice_hours}\n{lab_hours}\n{kurs_hours}'  # строка со значениями
        temp_df = pd.DataFrame([{'Тема': margint_text, 'Количество_часов': value_text}])
        df = pd.concat([df, temp_df], ignore_index=True)  # добаляем итоговую строку
        main_df = pd.concat([main_df, df], ignore_index=True)  # добавляем в основной датафрейм

    main_df.insert(0, 'Номер', np.nan)  # добавляем колонку с номерами занятий

    main_df['Содержание'] = main_df['Содержание'].fillna('Пусто')  # заменяем наны на пусто

    count = 0  # счетчик
    for idx, row in enumerate(main_df.itertuples()):
        if (row[5] == 'Пусто') | ('Итого часов' in row[5]):
            main_df.iloc[idx, 0] = ''
        else:
            count += 1
            main_df.iloc[idx, 0] = count

    # очищаем от пустых символов и строки Пусто
    main_df['Курс_семестр'] = main_df['Курс_семестр'].fillna('Пусто')
    main_df['Раздел'] = main_df['Раздел'].fillna('Пусто')

    main_df['Курс_семестр'] = main_df['Курс_семестр'].replace('Пусто', '')
    main_df['Тема'] = main_df['Тема'].replace('Пусто', '')
    main_df['Раздел'] = main_df['Раздел'].replace('Пусто', '')
    main_df['Содержание'] = main_df['Содержание'].replace('Пусто', '')

    main_df['Вид_занятия'] = main_df['Вид_занятия'].fillna('')

    main_df['Количество_часов'] = main_df['Количество_часов'].apply(convert_to_int)
    main_df['Количество_часов'] = main_df['Количество_часов'].fillna('')
    main_df['Практика'] = main_df['Практика'].fillna(0)
    main_df['Практика'] = main_df['Практика'].astype(int, errors='ignore')
    main_df['Практика'] = main_df['Практика'].apply(lambda x: '' if x == 0 else x)

    main_df['СРС'] = main_df['СРС'].fillna(0)
    main_df['СРС'] = main_df['СРС'].astype(int, errors='ignore')
    main_df['СРС'] = main_df['СРС'].apply(lambda x: '' if x == 0 else x)
    main_df['Содержание'] = main_df['Курс_семестр'] + main_df['Раздел'] + main_df['Тема'] + main_df['Содержание']
    main_df.drop(columns=['Курс_семестр', 'Раздел', 'Тема'], inplace=True)

    return (main_df,dct_all_sum_result) # возвращаем кортеж




def processing_mdk(data_pm) -> dict:
    """
    Функция для обработки листов с названием МДК
    :param data_pm: путь к файлу
    :return: список датафреймов
    """
    # Получаем список листов

    dct_mdk = dict()
    wb = openpyxl.load_workbook(data_pm,read_only=True) # получаем названия листов содержащих МДК
    for sheet_name in wb.sheetnames:
        if 'МДК' in sheet_name:
            name_mdk = str(wb[sheet_name]['D1'].value) # получаем название МДК, делаем строковыми на случай нан
            if 'МДК' in name_mdk:
                temp_mdk_df,temp_dct_result = extract_data_mdk(data_pm,sheet_name) # извлекаем данные из датафрейма
                dct_mdk[name_mdk] = {'Итог':temp_dct_result,'Данные':temp_mdk_df}
    wb.close()
    return dct_mdk





def create_rp_pp(template_pm: str, data_pm: str, end_folder: str):
    """
    Функция для создания программ профессиональных модулей
    :param template_pm: шаблон профмодуля
    :param data_pm: таблица Excel
    :param end_folder: конечная папка
    :return:
    """
    # названия листов
    desc_rp = 'Описание ПМ'
    pers_result = 'Лич_результаты'
    volume_pm = 'Объем ПМ'
    volume_all_mdk = 'Объем МДК'
    kp = 'Тематика КП(КР)'
    up = 'УП'
    pp = 'ПП'
    mto = 'МТО'
    main_publ = 'ОИ'
    second_publ = 'ДИ'
    ii_publ = 'ИИ'
    control = 'Контроль'
    pk = 'ПК'
    ok = 'ОК'
    fgos = 'Данные ФГОС'
    try:
        etalon_cols_lst = [desc_rp, pers_result, volume_pm, volume_all_mdk, kp, up, pp, mto, main_publ, second_publ,
                           ii_publ, control, pk, ok, fgos]
        etalon_cols = set(etalon_cols_lst)
        temp_wb = openpyxl.load_workbook(data_pm, read_only=True)
        file_cols = set(temp_wb.sheetnames)
        diff_cols = etalon_cols - file_cols
        temp_wb.close()
        if len(diff_cols) != 0:
            raise DiffSheet
        # Обрабатываем лист Описание ПМ
        df_desc_rp = pd.read_excel(data_pm, sheet_name=desc_rp, nrows=1, usecols='A:K')  # загружаем датафрейм
        df_desc_rp.fillna('НЕ ЗАПОЛНЕНО !!!', inplace=True)  # заполняем не заполненные разделы
        df_desc_rp.columns = ['Тип_программы', 'Название_модуля', 'Цикл', 'ВПД','Перечень', 'Код_специальность_профессия',
                              'Год', 'Разработчик', 'Методист', 'Название_ПЦК', 'Пред_ПЦК']

        # Создаем переменные для ФИО утверждающих
        df_accept_fio = pd.read_excel(data_pm, sheet_name=desc_rp, nrows=2, usecols='L:M')  # загружаем датафрейм
        accept_UR = df_accept_fio.iloc[0,1]
        accept_PR = df_accept_fio.iloc[1,1]

        # Обрабатываем лист Лич_результаты

        df_pers_result = pd.read_excel(data_pm, sheet_name=pers_result, usecols='A')
        df_pers_result.dropna(inplace=True)  # удаляем пустые строки
        df_pers_result.columns = ['Описание']
        df_pers_result['Код'] = df_pers_result['Описание'].apply(extract_lr)
        df_pers_result['Результат'] = df_pers_result['Описание'].apply(extract_descr_lr)

        # # Обрабатываем лист Объем ПМ
        df_volume_pm = pd.read_excel(data_pm,sheet_name=volume_pm,usecols='A:B')
        df_volume_pm.dropna(inplace=True) # удаляем пустые строки
        df_volume_pm.columns = ['Наименование', 'Объем']
        df_volume_pm.set_index('Наименование',inplace=True) # делаем индексом первую колонку
        _dct_df_volume_pm = df_volume_pm.to_dict('dict') # превращаем в словарь
        dct_df_volume_pm = _dct_df_volume_pm['Объем']

        """
            Обрабатываем лист Объем МДК
            """
        df_volume_all_mdk = pd.read_excel(data_pm,sheet_name=volume_all_mdk,usecols='A:J')
        df_volume_all_mdk.columns = ['Наименование','Всего','Прак_под','Обяз','Прак_зан','КР','СРС','УП','ПП','КА']
        df_volume_all_mdk.dropna(inplace=True,thresh=1) # удаляем пустые строки
        df_volume_all_mdk.fillna(0,inplace=True)  # заполняем наны
        df_volume_all_mdk.iloc[:,1:] = df_volume_all_mdk.iloc[:,1:].applymap(lambda x: int(x) if isinstance(x,(int,float)) else 0) # приводим к инту
        sum_row = df_volume_all_mdk.sum() # получаем строку общей суммы
        df_volume_all_mdk.loc['Сумма'] = sum_row # добавляем строку в датафрейм
        df_volume_all_mdk.at['Сумма','Наименование'] = 'Итого'
        df_volume_all_mdk = df_volume_all_mdk.astype(int,errors='ignore') # делай интовыми
        df_volume_all_mdk = df_volume_all_mdk.applymap(lambda x:'' if x ==0 else x)

        # # Открываем файл
        # wb = openpyxl.load_workbook(data_pm, read_only=True)
        # target_value = 'итог'
        #
        # # Поиск значения в выбранном столбце
        # column_number = 1  # Номер столбца, в котором ищем значение (например, столбец A)
        # target_row = None  # Номер строки с искомым значением
        #
        # for row in wb['Объем УД'].iter_rows(min_row=1, min_col=column_number, max_col=column_number):
        #     cell_value = row[0].value
        #     if target_value in str(cell_value).lower():
        #         target_row = row[0].row
        #         break
        #
        # if not target_row:
        #     # если не находим строку в которой есть слово итог то выдаем исключение
        #     messagebox.showerror('Диана Создание рабочих программ',
        #                          'Не найдена строка с названием Итоговая аттестация в листе Объем УД')
        #
        # wb.close()  # закрываем файл
        #
        # # если значение найдено то считываем нужное количество строк и  7 колонок
        # df_structure = pd.read_excel(data_pm, sheet_name=structure, nrows=target_row,
        #                              usecols='A:C', dtype=str)
        #
        # df_structure.iloc[:-1, 1:3] = df_structure.iloc[:-1, 1:3].applymap(convert_to_int)
        # df_structure.columns = ['Вид', 'Всего', 'Практика']
        # df_structure.fillna('', inplace=True)
        # max_load = df_structure.loc[0, 'Всего']  # максимальная учебная нагрузка
        # mand_load = df_structure.loc[1, 'Всего']  # обязательная нагрузка
        # practice_load = df_structure.loc[1, 'Практика']  # практическая нагрузка
        #
        # sam_df = df_structure[
        #     df_structure['Вид'] == 'Самостоятельная работа обучающегося (всего)']  # получаем часы самостоятельной работы
        # if len(sam_df) == 0:
        #     messagebox.showerror('Диана Создание рабочих программ',
        #                          'Проверьте наличие строки Самостоятельная работа обучающегося (всего) в листе Объем УД')
        # sam_load = sam_df.iloc[0, 1]

        """
            Обрабатываем листы с МДК
            """

        _dct_mdk_df = processing_mdk(data_pm) # получам словарь где ключ это название МДК а значение это словарь с данными и итогами по подсчету этих данных
        dct_mdk_df ={mdk:value['Данные'] for mdk,value in _dct_mdk_df.items()} # создаем словарь извлекая датафрейм
        _dct_mdk_data ={mdk:value['Итог'] for mdk,value in _dct_mdk_df.items()} # создаем словарь извлекая словарь с данными
        dct_mdk_data = dict() # считаем общую сумму
        for name,dct in _dct_mdk_data.items():
            for key,value in dct.items():
                if key not in dct_mdk_data:
                    dct_mdk_data[key] = value
                else:
                    dct_mdk_data[key] += value

        """Обрабатываем лист ПК
               """
        df_pk = pd.read_excel(data_pm, sheet_name=pk, usecols='A:C')
        df_pk.dropna(inplace=True, thresh=1)  # удаляем пустые строки
        # Обработка ПК
        lst_pk = df_pk['Наименование ПК'].dropna().tolist()
        lst_pk = processing_punctuation_end_string(lst_pk, ';\n', '- ', '.')
        df_pk.fillna('',inplace=True)
        df_pk.columns = ['Наименование','Показатель','Форма']

        # Обработка ОК
        df_ok = pd.read_excel(data_pm, sheet_name=ok, usecols='A:C')
        df_ok.dropna(inplace=True, thresh=1)  # удаляем пустые строки
        lst_ok = df_ok['Наименование ОК'].dropna().tolist()
        lst_ok = processing_punctuation_end_string(lst_ok, ';\n', '- ', '.')

        # Разворачиваем ОК в две колонки
        df_flat_ok = df_ok['Наименование ОК'].to_frame()
        df_ok.fillna('',inplace=True) # заполняем пустыми пробелами наны
        df_ok.columns = ['Наименование', 'Показатель', 'Форма']

        df_flat_ok.dropna(inplace=True)

        df_flat_ok.columns = ['Описание']
        df_flat_ok['Код'] = df_flat_ok['Описание'].apply(extract_lr)
        df_flat_ok['Результат'] = df_flat_ok['Описание'].apply(extract_descr_lr)

        """
            Обрабатываем лист Контроль
            """
        df_control = pd.read_excel(data_pm, sheet_name=control, usecols='A:B')
        df_control.dropna(inplace=True, thresh=1)  # удаляем пустые строки
        df_control.columns = ['Результаты_обучения', 'Контроль_обучения']
        _lst_result_educ = df_control['Результаты_обучения'].dropna().tolist()  # создаем список
        control_set = {'Знания:','Умения:','Практический опыт:'}

        if not control_set.issubset(set(_lst_result_educ)): # проверяем наличие нужных слов
            raise ControlWord_PM

        border_divide = _lst_result_educ.index('Знания:')  # граница разделения знания
        border_divide_second = _lst_result_educ.index('Практический опыт:') # граница разделения Опыта
        lst_skill = _lst_result_educ[1:border_divide]  # получаем список умений
        lst_knowledge = _lst_result_educ[border_divide + 1:border_divide_second]  # получаем список знаний
        lst_prac_exp = _lst_result_educ[border_divide_second+1:]

        lst_skill = processing_punctuation_end_string(lst_skill, ';\n', '- ', '.')  # форматируем выходную строку
        lst_knowledge = processing_punctuation_end_string(lst_knowledge, ';\n', '- ', '.')  # форматируем выходную строку
        lst_prac_exp = processing_punctuation_end_string(lst_prac_exp, ';\n', '- ', '.')  # форматируем выходную строку

        df_control.fillna('', inplace=True)
        # создаем 2 датафрейма для умений и практики
        df_control_knowldege = df_control.iloc[:border_divide-1,:]
        df_control_prac = df_control.iloc[border_divide_second:,:]
        out_contol_df = pd.concat([df_control_prac,df_control_knowldege],ignore_index=True,axis=0)

        """
            Обрабатываем лист темы курсовых работ
                """
        df_kp = pd.read_excel(data_pm,sheet_name=kp,usecols='A')
        lst_kp = df_kp.iloc[:,0].dropna().tolist()
        lst_kp = processing_punctuation_end_string(lst_kp, ';\n', '- ', '.')  # форматируем выходную строку

        """
            Обрабатываем лист УП (Учебная практика)
            """
        df_up = pd.read_excel(data_pm, sheet_name=up, usecols='A:D')
        df_up.columns = ['Вид','Содержание','Объем','Прак_под']
        df_up.fillna(0,inplace=True)
        df_up = df_up.applymap(lambda x:int(x) if isinstance(x,float) else x)
        df_up = df_up.applymap(lambda x: '' if x ==0 else x)

        # делим датафрейм по разделам
        borders = df_up[
            df_up['Вид'].str.contains('Раздел')].index  # получаем индексы строк где есть слово семестр

        part_df_up = []  # список для хранения кусков датафрейма
        previos_border = -1
        # делим датафрем по границам
        for value_border in borders:
            part = df_up.iloc[previos_border:value_border]
            part_df_up.append(part)
            previos_border = value_border

        # добавляем последнюю часть
        last_part = df_up.iloc[borders[-1]:]
        part_df_up.append(last_part)
        part_df_up.pop(0)  # удаляем нулевой элемент так как он пустой
        df_up_short = pd.DataFrame(columns=['Вид','Объем','Прак_под']) # короткий датафрейм для объема
        df_up_content = pd.DataFrame(columns=['Вид','Содержание','Объем']) # короткий датафрейм для содержания
        for df in part_df_up:
            name_part = [value.strip() for value in df['Вид'].tolist() if 'Раздел' in value][0] # получаем название раздела
            volume_part = int(sum_column_any_value(df,'Объем'))
            prac_volume_part = int(sum_column_any_value(df,'Прак_под'))
            temp_df = pd.DataFrame(columns=['Вид','Объем','Прак_под'],
                                   data=[[name_part,volume_part,prac_volume_part]])
            df_up_short = pd.concat([df_up_short,temp_df],axis=0,ignore_index=True)

            # Суммируем
            content_temp_df = df[['Вид','Содержание','Объем']].copy()
            content_temp_df.loc['Итого'] = int(sum_column_any_value(content_temp_df,'Объем'))
            # убираем лишние цифры
            content_temp_df.at['Итого','Вид'] = ''
            content_temp_df.at['Итого','Содержание'] = 'Итого'
            df_up_content = pd.concat([df_up_content,content_temp_df],ignore_index=True,axis=0)

        sum_volume_up = sum_column_any_value(df_up,'Объем')
        sum_practice_up = sum_column_any_value(df_up,'Прак_под')

        theme_up_df = (df_up['Вид'] + df_up['Содержание']).to_frame()
        lst_theme_up = theme_up_df.iloc[:,0].dropna().tolist()
        lst_theme_up = processing_punctuation_end_string(lst_theme_up, ';\n', '- ', '.')  # форматируем выходную строку

        """
            Обрабатываем лист ПП (Производственная практика)
            """
        df_pp = pd.read_excel(data_pm, sheet_name=pp, usecols='A:D')
        df_pp.columns = ['Вид','Содержание','Объем','Прак_под']
        df_pp.fillna(0,inplace=True)
        df_pp = df_pp.applymap(lambda x:int(x) if isinstance(x,float) else x)
        df_pp = df_pp.applymap(lambda x: '' if x ==0 else x)

        # делим датафрейм по разделам
        borders = df_pp[
            df_pp['Вид'].str.contains('Раздел')].index  # получаем индексы строк где есть слово семестр

        part_df_pp = []  # список для хранения кусков датафрейма
        previos_border = -1
        # делим датафрем по границам
        for value_border in borders:
            part = df_pp.iloc[previos_border:value_border]
            part_df_pp.append(part)
            previos_border = value_border

        # добавляем последнюю часть
        last_part = df_pp.iloc[borders[-1]:]
        part_df_pp.append(last_part)
        part_df_pp.pop(0)  # удаляем нулевой элемент так как он пустой
        df_pp_short = pd.DataFrame(columns=['Вид', 'Объем', 'Прак_под'])  # короткий датафрейм для объема
        df_pp_content = pd.DataFrame(columns=['Вид', 'Содержание', 'Объем'])  # короткий датафрейм для содержания
        for df in part_df_pp:
            name_part = [value.strip() for value in df['Вид'].tolist() if 'Раздел' in value][0]  # получаем название раздела
            volume_part = int(sum_column_any_value(df, 'Объем'))
            prac_volume_part = int(sum_column_any_value(df, 'Прак_под'))
            temp_df = pd.DataFrame(columns=['Вид', 'Объем', 'Прак_под'],
                                   data=[[name_part, volume_part, prac_volume_part]])
            df_pp_short = pd.concat([df_pp_short, temp_df], axis=0, ignore_index=True)

            # Суммируем
            content_temp_df = df[['Вид', 'Содержание', 'Объем']].copy()
            content_temp_df.loc['Итого'] = int(sum_column_any_value(content_temp_df, 'Объем'))
            # убираем лишние цифры
            content_temp_df.at['Итого', 'Вид'] = ''
            content_temp_df.at['Итого', 'Содержание'] = 'Итого'
            df_pp_content = pd.concat([df_pp_content, content_temp_df], ignore_index=True, axis=0)

        sum_volume_pp = sum_column_any_value(df_pp, 'Объем')
        sum_practice_pp = sum_column_any_value(df_pp, 'Прак_под')

        theme_pp_df = (df_pp['Вид'] + df_pp['Содержание']).to_frame()
        lst_theme_pp = theme_pp_df.iloc[:,0].dropna().tolist()
        lst_theme_pp = processing_punctuation_end_string(lst_theme_pp, ';\n', '- ', '.')  # форматируем выходную строку

        """
            Обрабатываем лист МТО
            """
        df_mto = pd.read_excel(data_pm, sheet_name=mto, usecols='A:G')
        df_mto.dropna(inplace=True, thresh=1)  # удаляем пустые строки
        name_kab = df_mto['Наименование_учебного_кабинета'].dropna().tolist()  #
        name_kab = ','.join(name_kab)  # получаем все названия которые есть в колонке Наименование учебного кабинета

        name_lab = df_mto['Наименование_лаборатории'].dropna().tolist()
        if len(name_lab) == 0:
            name_lab = []
        else:
            name_lab = ','.join(name_lab)

        name_work = df_mto['Наименование_мастерской'].dropna().tolist()
        if len(name_lab) == 0:
            name_work = []
        else:
            name_work = ','.join(name_work)

        # Списки кабинета и средств обучения
        lst_obor_cab = df_mto[
            'Оборудование_учебного_кабинета'].dropna().tolist()  # создаем список удаляя все незаполненные ячейки
        if len(lst_obor_cab) == 0:
            lst_obor_cab = ['На листе МТО НЕ заполнено оборудование учебного кабинета !!!']
        obor_cab = processing_punctuation_end_string(lst_obor_cab, ';\n', '- ',
                                                     '.')  # обрабатываем знаки пунктуации для каждой строки

        lst_tecn_educ = df_mto[
            'Технические_средства_обучения'].dropna().tolist()  # создаем список удаляя все незаполненные ячейки
        if len(lst_tecn_educ) == 0:
            lst_tecn_educ = ['На листе МТО НЕ заполнены технические средства обучения !!!']
        tecn_educ = processing_punctuation_end_string(lst_tecn_educ, ';\n', '- ', '.')

        # Оборудование лаборатории
        lst_obor_labor = df_mto[
            'Оборудование_лаборатории'].dropna().tolist()  # создаем список удаляя все незаполненные ячейки
        if len(lst_obor_labor) == 0 and name_lab:
            obor_labor = ['На листе МТО НЕ заполнено оборудование лаборатории !!!']
        elif len(lst_obor_labor) != 0 and not name_lab:
            obor_labor = ['На листе МТО заполнено оборудование лаборатории но не заполнено наименование лаборатории !!!']
        else:
            obor_labor = processing_punctuation_end_string(lst_obor_labor, ';\n', '- ',
                                                           '.')  # обрабатываем знаки пунктуации для каждой строки

        # Оборудование мастерской
        lst_tecn_work = df_mto[
            'Оборудование_мастерской'].dropna().tolist()  # создаем список удаляя все незаполненные ячейки
        if len(lst_tecn_work) == 0:
            lst_tecn_work = ['На листе МТО НЕ заполнены технические средства обучения !!!']
        tecn_work = processing_punctuation_end_string(lst_tecn_work, ';\n', '- ', '.')

        """
            Обрабатываем лист Основные источники
            """

        df_main_publ = pd.read_excel(data_pm, sheet_name=main_publ, usecols='A:G')
        if df_main_publ.shape[0] != 0:
            df_main_publ.dropna(inplace=True, thresh=1)  # удаляем пустые строки
            df_main_publ.fillna('Не заполнено !!!', inplace=True)
            df_main_publ = df_main_publ.applymap(str)  # приводим к строковому виду
            df_main_publ = df_main_publ.applymap(str.strip)  # очищаем от пробелов в начале и конце

            df_main_publ['Основной_источник'] = df_main_publ.apply(processing_publ, axis=1)  # формируем строку
            df_main_publ.sort_values(by='Основной_источник', inplace=True)
            lst_main_source = df_main_publ['Основной_источник'].tolist()
        else:
            lst_main_source = 'Не заполнено'

        """
            Обрабатываем лист дополнительные источники
            """
        df_second_publ = pd.read_excel(data_pm, sheet_name=second_publ, usecols='A:G')
        if df_second_publ.shape[0] != 0:
            df_second_publ.dropna(inplace=True, thresh=1)  # удаляем пустые строки
            df_second_publ.fillna('Не заполнено !!!', inplace=True)
            df_second_publ = df_second_publ.applymap(str)  # приводим к строковому виду
            df_second_publ = df_second_publ.applymap(str.strip)  # очищаем от пробелов в начале и конце

            df_second_publ['Основной_источник'] = df_second_publ.apply(processing_publ, axis=1)  # формируем строку
            df_second_publ.sort_values(by='Основной_источник', inplace=True)
            lst_slave_source = df_second_publ['Основной_источник'].tolist()
        else:
            lst_slave_source = 'Не заполнено'

        """
            Обрабатываем лист интернет источники
            """
        df_ii_publ = pd.read_excel(data_pm, sheet_name=ii_publ, usecols='A:B')
        if df_ii_publ.shape[0] != 0:
            df_ii_publ.dropna(inplace=True, thresh=1)  # удаляем пустые строки

            df_ii_publ.sort_values(by='Название', inplace=True)

            lst_inet_source = insert_type_source(df_ii_publ.copy())
        else:
            lst_inet_source = 'Не заполнено'

        """
            Обрабатываем лист ФГОС
            """
        df_fgos = pd.read_excel(data_pm,sheet_name=fgos,usecols='A:C')
        df_fgos.dropna(inplace=True, thresh=1)  # удаляем пустые строки
        df_fgos.fillna('Не заполнено',inplace=True)

        doc = DocxTemplate(template_pm)

        # Конвертируем датафрейм с описанием программы в список словарей и добавляем туда нужные элементы
        data_program = df_desc_rp.to_dict('records')
        context = data_program[0]
        context['ФИО_УР'] = accept_UR # Переменные для ФИО утверждающих УР и ПР
        context['ФИО_ПР'] = accept_PR

        context['Лич_результаты'] = df_pers_result.to_dict('records')  # добаввляем лист личностных результатов
        context['Объем_ПМ'] = df_volume_pm.to_dict('records')  # объем ПМ
        # context['Учебная_работа'] = df_structure.to_dict('records')
        context['Контроль_оценка'] = df_control.to_dict('records')
        context['Практика_умение'] = out_contol_df.to_dict('records')
        context['Знать'] = lst_knowledge
        context['Уметь'] = lst_skill
        context['Прак_опыт'] = lst_prac_exp

        # Объем МДК
        context['Объем_МДК'] = df_volume_all_mdk.to_dict('records')

        # добавляем единичные переменные
        context['Всего'] = dct_df_volume_pm['Всего часов']
        context['Макс_уч_нагр'] = dct_df_volume_pm['Максимальной учебной нагрузки обучающегося']
        context['Обяз_ауд_нагр'] = dct_df_volume_pm['Обязательной аудиторной нагрузки обучающегося']
        context['КР'] = dct_df_volume_pm['курсовой проект (работа)']
        context['Прак_подг'] = dct_df_volume_pm['на практическую подготовку']
        context['СРС'] = dct_df_volume_pm['самостоятельная работа обучающегося']
        context['Консул'] = dct_df_volume_pm['консультации']
        context['Пром_атт'] = dct_df_volume_pm['промежуточная аттестация']
        context['Экзамен_квал'] = dct_df_volume_pm['экзамен (квалификационный)']
        context['Объем_УП'] = sum_volume_up # берем значения из соответсвующего листа
        context['Объем_УП_прак_под'] = sum_practice_up
        context['Объем_ПП'] = sum_volume_pp
        context['Объем_ПП_прак_под'] = sum_practice_pp
        context['Атт_УП'] = dct_df_volume_pm['итоговая аттестация УП в форме']
        context['Атт_ПП'] = dct_df_volume_pm['итоговая аттестация ПП в форме']
        context['Квалификация'] = df_volume_pm.iloc[-1,0] # получаем значение ячейки на последней строке

        context['Темы_КР'] = lst_kp # список тем курсовых работ
        #
        context['Темы_УП'] = lst_theme_up
        context['Темы_ПП'] = lst_theme_pp

        context['УП_разд_объем'] = df_up_short.to_dict('records') # делаем таблицу УП
        context['УП_содер_объем'] = df_up_content.to_dict('records') # делаем таблицу УП

        context['ПП_разд_объем'] = df_pp_short.to_dict('records') # делаем таблицу УП
        context['ПП_содер_объем'] = df_pp_content.to_dict('records') # делаем таблицу УП

        # #лист МТО
        context['Учебный_кабинет'] = name_kab
        context['Лаборатория'] = name_lab
        context['Мастерская'] = name_work
        context['Оборудование_кабинета'] = obor_cab
        context['Средства_обучения'] = tecn_educ
        context['Оборудование_лаборатории'] = obor_labor
        context['Оборудование_мастерской'] = tecn_work
        # лист Учебные издания
        context['Основные_источники'] = lst_main_source
        context['Дополнительные_источники'] = lst_slave_source
        context['Интернет_источники'] = lst_inet_source
        # Листы данные ОК,ПК
        context['ОК'] = lst_ok
        context['ПК'] = lst_pk
        context['Контроль_ПК'] = df_pk.to_dict('records')
        context['Контроль_ОК'] = df_ok.to_dict('records')

        context['Общ_компетенции'] = df_flat_ok.to_dict('records')

        # Переменные ФГОС
        # даты
        context['ФГОС_у_дата'] = df_fgos.at[0,'Дата']
        context['ФГОС_з_дата'] = df_fgos.at[1,'Дата']
        context['Прог_з_дата'] = df_fgos.at[2,'Дата']

        # номера
        context['ФГОС_у_номер'] = df_fgos.at[0,'Номер документа']
        context['ФГОС_з_номер'] = df_fgos.at[1,'Номер документа']
        context['Прог_з_номер'] = df_fgos.at[2,'Номер документа']

        for idx,tpl_dct in enumerate(dct_mdk_df.items(),start=1):
            key,value = tpl_dct # распаковываем кортеж
            # делаем переменные для названий МДК
            name_var = f'МДК_{idx}'
            context[name_var] = key
            # создаем переменные датафреймов
            name_mdk_table = f'План_МДК_{idx}'
            context[name_mdk_table] = value.to_dict('records')
        # Создаем документ
        doc.render(context)
        # сохраняем документ
        # название программы
        name_rp = df_desc_rp['Название_модуля'].tolist()[0]
        # name_rp = 'Тест'
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        doc.save(f'{end_folder}/РП Производственная практика {name_rp[:40]} {current_time}.docx')
    except DiffSheet:
        messagebox.showerror('Диана Создание рабочих программ',
                             f'В таблице не найдены листы {diff_cols},\n'
                             f'В таблице должны быть листы {etalon_cols_lst}\n'
                             f'Возможно вы используете шаблон РП для ООД а не шаблон РП для УД')
    except ControlWord_PM:
        messagebox.showerror('Диана Создание рабочих программ',
                             'На листе Контроль в первой колонке должно быть указаны слова\n'
                             'Умения: , Знания: , Практический опыт:\n'
                             'Посмотрите пример в исходном шаблоне')
    except KeyError as e:
        messagebox.showerror('Диана Создание рабочих программ',
                             f'В таблице не найдена колонка с названием {e.args}!\nПроверьте написание названия колонки')
    except ValueError as e:
        messagebox.showerror('Диана Создание рабочих программ',
                             f'В таблице не найден лист,колонка или значение {e.args}!\nПроверьте написание названий')

    except FileNotFoundError:
        messagebox.showerror('Диана Создание рабочих программ',
                             f'Перенесите файлы которые вы хотите обработать в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам')

    except exceptions.TemplateSyntaxError:
        messagebox.showerror('Диана Создание рабочих программ',
                             f'Ошибка в оформлении вставляемых значений в шаблоне\n'
                             f'Проверьте свой шаблон на наличие следующих ошибок:\n'
                             f'1) Вставляемые значения должны быть оформлены двойными фигурными скобками\n'
                             f'{{{{Вставляемое_значение}}}}\n'
                             f'2) В названии колонки в таблице откуда берутся данные - есть пробелы,цифры,знаки пунктуации и т.п.\n'
                             f'в названии колонки должны быть только буквы и нижнее подчеркивание.\n'
                             f'{{{{Дата_рождения}}}}')
    else:
        messagebox.showinfo('Диана Создание рабочих программ', 'Данные успешно обработаны')


# messagebox.showerror('Диана Создание рабочих программ',
#                      'На листе Контроль в первой колонке должно быть указаны слова\n'
#                      'Умения: , Знания: , Практический опыт:\n'
#                      'Посмотрите пример в исходном шаблоне')

# jinja2.exceptions.UndefinedError: 'Интернет' is undefined неправильная запись в шаблоне

if __name__ == '__main__':
    template_pm_main = 'data/Шаблон автозаполнения ПП.docx'
    data_pm_main = 'data/Таблица для ПМ,УП,ПП.xlsx'
    end_folder_main = 'data'

    create_rp_pp(template_pm_main, data_pm_main, end_folder_main)
    print('Lindy Booth')

