"""
скрипт для создания рабочих программ преддипломной практики
"""
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from docxtpl import DocxTemplate
import string
import time
import re
from tkinter import messagebox
from jinja2 import exceptions
pd.options.mode.chained_assignment = None  # default='warn'
pd.set_option('display.max_columns', None)  # Отображать все столбцы
pd.set_option('display.expand_frame_repr', False)  # Не переносить строки
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.filterwarnings('ignore', category=FutureWarning, module='openpyxl')
warnings.filterwarnings("ignore", category=DeprecationWarning)


class DiffSheet(Exception):
    """
    Исключение для случаев когда отсутствуют нужные колонки в файле
    """
    pass



def write_df_to_excel(dct_df:dict,write_index:bool)->openpyxl.Workbook:
    """
    Функция для записи датафрейма в файл Excel
    :param dct_df: словарь где ключе это название создаваемого листа а значение датафрейм который нужно записать
    :param write_index: нужно ли записывать индекс датафрейма True or False
    :return: объект Workbook с записанными датафреймами
    """
    wb = openpyxl.Workbook() # создаем файл
    count_index = 0 # счетчик индексов создаваемых листов
    for name_sheet,df in dct_df.items():
        wb.create_sheet(title=name_sheet,index=count_index) # создаем лист
        # записываем данные в лист
        for row in dataframe_to_rows(df,index=write_index,header=True):
            wb[name_sheet].append(row)
        # ширина по содержимому
        # сохраняем по ширине колонок
        for column in wb[name_sheet].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[name_sheet].column_dimensions[column_name].width = adjusted_width
        count_index += 1
    # удаляем лишний лист
    if len(wb.sheetnames) >= 2 and 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    return wb
def convert_to_int(cell):
    """
    Метод для проверки значения ячейки
    :param cell: значение ячейки
    :return: число в формате int
    """
    if cell is np.nan:
        return ''
    try:
        value = float(cell)
        return int(value)
    except:
        return cell


def extract_lr(cell):
    """
    Функция для создания 2 списков из данных одной колонки
    первая колонка это код личностного результата а вторая колонка  это описание
    извлечение будет с помощью регулярок
    :param cell: ячейка датафрейма
    :return:
    """
    value = str(cell)  # делаем строковй
    result = re.split(r'(\d+\.\s*?)', value)
    result = [value for value in result if value]  # убираем пустые значения
    if len(result) >= 3:
        end_lr = result[1].strip()
        end_lr = end_lr.rstrip(string.punctuation)  # очищаем от точки в конце
        return f'{result[0]}{end_lr}'
    else:
        lr = result[0].strip()
        lr = lr.rstrip(string.punctuation)
        return lr


def extract_descr_lr(cell):
    """
    Функция для создания 2 списков из данных одной колонки
    первая колонка это код личностного результата а вторая колонка  это описание
    извлечение будет с помощью регулярок
    :param cell: ячейка
    :return:
    """
    value = str(cell)  # делаем строковй
    result = re.split(r'(\d+\.\s*?)', value)
    result = [value for value in result if value]
    if len(result) >= 3:
        descr_lr = result[2].strip()
        descr_lr = descr_lr.rstrip(string.punctuation)  # очищаем от точки в конце
        return f'{descr_lr}.'
    else:
        return ''


def processing_punctuation_end_string(lst_phrase: list, sep_string: str, sep_begin: str, sep_end: str) -> str:
    """
    Очистка каждого элемента списка от знаков пунктации в конце,
    добавление разделитея, добавление точки в конце
    :param lst_phrase: список элементов
    :param sep_string: разделитель между элементами списка
    :param sep_begin: начальный разделитель
    :param sep_end: знак пунктуации в конце
    :return: строку с разделителями и переносами строки
    """
    if len(lst_phrase) == 0:
        return ''
    lst_phrase = list(map(str, lst_phrase))
    temp_lst = list(map(lambda x: sep_begin + x, lst_phrase))
    temp_lst = list(map(lambda x: x.strip(), temp_lst))  # очищаем от прбельных символов в начале и конце
    temp_lst = list(map(lambda x: x.rstrip(string.punctuation), temp_lst))  # очищаем от знаков пунктуации
    temp_lst[-1] = temp_lst[-1] + sep_end  # добавляем конечный знак пунктуации
    temp_str = f'{sep_string}'.join(temp_lst)  # создаем строку с разделителями
    return temp_str


def insert_type_source(df: pd.DataFrame) -> list:
    """
    Вставка в строку слов [Электронный ресурс] Форма доступа:
    :param lst_phrase:датафрейм
    :return: список измененных строк
    """
    out_lst = []  # список для хранения строк
    for row in df.itertuples():
        name = str(row[1])
        url_ii = str(row[2])
        name = name.strip()  # очищаем от пробельных символов
        name = name.rstrip(string.punctuation)  # очищаем от знаков препинания
        name = name.strip()
        url_ii = url_ii.strip()  # очищаем от пробельных символов
        url_ii = url_ii.rstrip(string.punctuation)  # очищаем от знаков препинания
        url_ii = url_ii.strip()
        temp_str = f'{name} [Электронный ресурс] Форма доступа:{url_ii}'
        out_lst.append(temp_str)

    return out_lst


def processing_publ(row):
    """
    Функция для генерации строки с литературой в нужном формате
    :param row: строка датафрейма
    :return: сумма строк датафрейма в нужном формате
    """
    author = row[0]  # автор(ы)
    name_book = row[1]  # название
    full_city = row[2]  # полное название города
    short_city = row[3]  # краткое название города
    publ_house = row[4]  # издательство
    year = row[5]  # год издания
    quan_pages = row[6]  # число страниц
    author = author.rstrip(string.punctuation)  # очищаем от символа пунктуации в конце
    name_book = name_book.rstrip(string.punctuation)  # очищаем от символа пунктуации в конце
    short_city = short_city.rstrip(string.punctuation)  # очищаем от символа пунктуации в конце
    publ_house = publ_house.rstrip(string.punctuation)  # очищаем от символа пунктуации в конце
    # извлекаем год
    result = re.search(r'\d{4}', year)
    if result:
        clean_year = result.group()
    else:
        clean_year = 'Неправильно заполнен год издания, введите год в формате 4 цифры без букв'

    # извлекаем количество страниц
    result = re.search(r'\d+', quan_pages)
    if result:
        clean_quan_pages = result.group()
    else:
        clean_quan_pages = 'Неправильно заполнено количество страниц, введите количество в виде числа без букв'
    # Формируем итоговую строку
    out_str = f'{author}. {name_book}.- {short_city}.: {publ_house}, {clean_year}.- {clean_quan_pages} c.'

    return out_str


def sum_column_any_value(df:pd.DataFrame,name_column:str):
    """
    Суммирование колонки с разными типами значений в том числе строковыми.
    """
    lst_value = df[name_column].dropna().tolist()
    sum_value = [value for value in lst_value if isinstance(value,(int,float))] # отбираем только числа
    return sum(sum_value) # возвращаем сумму

def create_pred_dip_prac(template_pm: str, data_pred_prac: str, end_folder: str):
    """
    Функция для создания программ профессиональных модулей
    :param template_pm: шаблон профмодуля
    :param data_pred_prac: таблица Excel
    :param end_folder: конечная папка
    :return:
    """
    # названия листов
    desc_pred_prac = 'Описание Пред_дип_практики'
    content_pred_prac = 'Содержание пред_дип_практики'
    pers_result = 'Лич_результаты'
    # mto = 'МТО'
    main_publ = 'ОИ'
    second_publ = 'ДИ'
    ii_publ = 'ИИ'
    pk = 'ПК'
    ok = 'ОК,ВД,ВПД'
    # vpd = 'ВПД'
    po = 'Практический опыт'
    # skills = 'Умения'
    fgos = 'Данные ФГОС'
    try:
        # etalon_cols_lst = [desc_pred_prac,content_pred_prac,pers_result,vpd,po,skills,mto,main_publ,second_publ,ii_publ,pk,ok,fgos]
        etalon_cols_lst = [desc_pred_prac,content_pred_prac,pers_result,po,main_publ,second_publ,ii_publ,pk,ok,fgos]
        etalon_cols = set(etalon_cols_lst)
        temp_wb = openpyxl.load_workbook(data_pred_prac, read_only=True)
        file_cols = set(temp_wb.sheetnames)
        diff_cols = etalon_cols - file_cols
        temp_wb.close()
        if len(diff_cols) != 0:
            raise DiffSheet

        # Обрабатываем лист Описание ПМ
        df_desc_rp = pd.read_excel(data_pred_prac, sheet_name=desc_pred_prac, nrows=1, usecols='A:H')  # загружаем датафрейм
        df_desc_rp.fillna('НЕ ЗАПОЛНЕНО !!!', inplace=True)  # заполняем не заполненные разделы
        df_desc_rp.columns = ['Тип_программы','Код_специальность_профессия','Форма_аттестации',
                              'Год', 'Разработчик', 'Методист', 'Название_ПЦК', 'Пред_ПЦК']

        form_att = df_desc_rp.at[0, 'Форма_аттестации'] # переменная для названия итоговой квалификации

        # Создаем переменные для ФИО утверждающих
        df_accept_fio = pd.read_excel(data_pred_prac, sheet_name=desc_pred_prac, nrows=1, usecols='I:J')  # загружаем датафрейм
        accept_PR = df_accept_fio.iloc[0,1]

        # Обрабатываем лист Лич_результаты

        df_pers_result = pd.read_excel(data_pred_prac, sheet_name=pers_result, usecols='A')
        df_pers_result.dropna(inplace=True)  # удаляем пустые строки
        df_pers_result.columns = ['Описание']
        df_pers_result['Код'] = df_pers_result['Описание'].apply(extract_lr)
        df_pers_result['Результат'] = df_pers_result['Описание'].apply(extract_descr_lr)


        """Обрабатываем лист ПК
               """
        df_pk = pd.read_excel(data_pred_prac, sheet_name=pk, usecols='A:C')
        df_pk.dropna(inplace=True, thresh=1)  # удаляем пустые строки
        # Обработка ПК
        lst_pk = df_pk['Наименование ПК'].dropna().tolist()
        lst_pk = processing_punctuation_end_string(lst_pk, ';\n', '- ', '.')
        df_pk.fillna('',inplace=True)
        df_pk.columns = ['Наименование','Показатель','Форма']

        # Обработка ОК
        df_ok_vd = pd.read_excel(data_pred_prac, sheet_name=ok, usecols='A:C')
        df_ok_vd.dropna(inplace=True, thresh=1)  # удаляем пустые строки
        df_ok_vd.columns = ['Наименование ОК','Вид деятельности','ВПД']
        lst_ok = df_ok_vd['Наименование ОК'].dropna().tolist() # список для ОК
        lst_ok = processing_punctuation_end_string(lst_ok, ';\n', '- ', '.')
        lst_vd = df_ok_vd['Вид деятельности'].dropna().tolist()  # список для Вид деятельности по ФГОС
        lst_vd = processing_punctuation_end_string(lst_vd, ';\n', '- ', '.')
        lst_vpd = df_ok_vd['ВПД'].dropna().tolist()  # список для Вид профессиональной деятельности
        lst_vpd = processing_punctuation_end_string(lst_vpd, ';\n', '- ', '.')

        """
            Обрабатываем лист ПП (Преддипломная практика)
            """
        df_pp = pd.read_excel(data_pred_prac, sheet_name=content_pred_prac, usecols='A:D')
        df_pp.columns = ['Вид', 'Содержание', 'Объем', 'Прак_под']
        df_pp.fillna(0, inplace=True)
        df_pp = df_pp.applymap(lambda x: int(x) if isinstance(x, float) else x)
        df_pp = df_pp.applymap(lambda x: '' if x == 0 else x)

        # делим датафрейм по разделам
        borders = df_pp[
            df_pp['Вид'].str.contains('Раздел')].index  # получаем индексы строк где есть слово семестр

        part_df_pp = []  # список для хранения кусков датафрейма
        previos_border = -1
        # делим датафрем по границам
        for value_border in borders:
            part = df_pp.iloc[previos_border:value_border]
            part_df_pp.append(part)
            previos_border = value_border

        # добавляем последнюю часть
        last_part = df_pp.iloc[borders[-1]:]
        part_df_pp.append(last_part)
        part_df_pp.pop(0)  # удаляем нулевой элемент так как он пустой
        df_pp_short = pd.DataFrame(columns=['Вид', 'Объем', 'Прак_под'])  # короткий датафрейм для объема
        df_pp_content = pd.DataFrame(columns=['Вид', 'Содержание', 'Объем'])  # короткий датафрейм для содержания
        for df in part_df_pp:
            name_part = [value.strip() for value in df['Вид'].tolist() if 'Раздел' in value][
                0]  # получаем название раздела
            volume_part = int(sum_column_any_value(df, 'Объем'))
            prac_volume_part = int(sum_column_any_value(df, 'Прак_под'))
            temp_df = pd.DataFrame(columns=['Вид', 'Объем', 'Прак_под'],
                                   data=[[name_part, volume_part, prac_volume_part]])
            df_pp_short = pd.concat([df_pp_short, temp_df], axis=0, ignore_index=True)

            # Суммируем
            content_temp_df = df[['Вид', 'Содержание', 'Объем']].copy()
            content_temp_df.loc['Итого'] = int(sum_column_any_value(content_temp_df, 'Объем'))
            # убираем лишние цифры
            content_temp_df.at['Итого', 'Вид'] = ''
            content_temp_df.at['Итого', 'Содержание'] = 'Итого'
            df_pp_content = pd.concat([df_pp_content, content_temp_df], ignore_index=True, axis=0)

        # Добавляем строку с названием аттестации
        df_pp_short.loc['Аттестация'] = f'Итоговая аттестация в форме {form_att}'
        df_pp_short.loc['Аттестация','Объем'] = ''
        df_pp_short.loc['Аттестация','Прак_под'] = ''

        df_pp_content.loc['Аттестация'] = f'Итоговая аттестация в форме {form_att}'
        df_pp_content.loc['Аттестация','Объем'] = ''
        df_pp_content.loc['Аттестация','Содержание'] = ''


        sum_volume_pp = sum_column_any_value(df_pp, 'Объем')
        sum_practice_pp = sum_column_any_value(df_pp, 'Прак_под')


        # """
        #     Обрабатываем лист МТО
        #     """
        # df_mto = pd.read_excel(data_pred_prac, sheet_name=mto, usecols='A:C')
        # df_mto.columns = ['Оборудование','Инструменты_приспособления','Средства_обучения']
        # df_mto.dropna(inplace=True, thresh=1)  # удаляем пустые строки
        # # Списки кабинета и средств обучения
        # lst_obor = df_mto[
        #     'Оборудование'].dropna().tolist()  # создаем список удаляя все незаполненные ячейки
        # if len(lst_obor) == 0:
        #     lst_obor = ['На листе МТО НЕ заполнена колонка Оборудование !!!']
        # lst_obor = processing_punctuation_end_string(lst_obor, ';\n', '- ',
        #                                              '.')  # обрабатываем знаки пунктуации для каждой строки
        #
        # lst_tecn = df_mto[
        #     'Инструменты_приспособления'].dropna().tolist()  # создаем список удаляя все незаполненные ячейки
        # if len(lst_tecn) == 0:
        #     lst_tecn = ['На листе МТО НЕ заполнена колонка Инструменты и приспособления !!!']
        # lst_tecn = processing_punctuation_end_string(lst_tecn, ';\n', '- ', '.')
        #
        # # Оборудование мастерской
        # lst_educ = df_mto[
        #     'Средства_обучения'].dropna().tolist()  # создаем список удаляя все незаполненные ячейки
        # if len(lst_educ) == 0:
        #     lst_educ = ['На листе МТО НЕ заполнена колонка Средства обучения !!!']
        # lst_educ = processing_punctuation_end_string(lst_educ, ';\n', '- ', '.')

        """
            Обрабатываем лист Основные источники
            """

        df_main_publ = pd.read_excel(data_pred_prac, sheet_name=main_publ, usecols='A:G')
        if df_main_publ.shape[0] != 0:
            df_main_publ.dropna(inplace=True, thresh=1)  # удаляем пустые строки
            df_main_publ.fillna('Не заполнено !!!', inplace=True)
            df_main_publ = df_main_publ.applymap(str)  # приводим к строковому виду
            df_main_publ = df_main_publ.applymap(str.strip)  # очищаем от пробелов в начале и конце

            df_main_publ['Основной_источник'] = df_main_publ.apply(processing_publ, axis=1)  # формируем строку
            df_main_publ.sort_values(by='Основной_источник', inplace=True)
            lst_main_source = df_main_publ['Основной_источник'].tolist()
        else:
            lst_main_source = 'Не заполнено'

        """
            Обрабатываем лист дополнительные источники
            """
        df_second_publ = pd.read_excel(data_pred_prac, sheet_name=second_publ, usecols='A:G')
        if df_second_publ.shape[0] != 0:
            df_second_publ.dropna(inplace=True, thresh=1)  # удаляем пустые строки
            df_second_publ.fillna('Не заполнено !!!', inplace=True)
            df_second_publ = df_second_publ.applymap(str)  # приводим к строковому виду
            df_second_publ = df_second_publ.applymap(str.strip)  # очищаем от пробелов в начале и конце

            df_second_publ['Основной_источник'] = df_second_publ.apply(processing_publ, axis=1)  # формируем строку
            df_second_publ.sort_values(by='Основной_источник', inplace=True)
            lst_slave_source = df_second_publ['Основной_источник'].tolist()
        else:
            lst_slave_source = 'Не заполнено'

        """
            Обрабатываем лист интернет источники
            """
        df_ii_publ = pd.read_excel(data_pred_prac, sheet_name=ii_publ, usecols='A:B')
        if df_ii_publ.shape[0] != 0:
            df_ii_publ.dropna(inplace=True, thresh=1)  # удаляем пустые строки

            df_ii_publ.sort_values(by='Название', inplace=True)

            lst_inet_source = insert_type_source(df_ii_publ.copy())
        else:
            lst_inet_source = 'Не заполнено'

        # Создаем датафрейм для таблицы Контроль

        control_df = pd.DataFrame(columns=['Наименование','Форма'])



        """
        Обрабатываем лист Практический опыт
        """
        df_prac = pd.read_excel(data_pred_prac, sheet_name=po, usecols='A:B')
        df_prac.dropna(inplace=True, thresh=1)  # удаляем пустые строки
        df_prac.fillna('', inplace=True)
        df_prac.columns = ['Наименование', 'Форма']
        # создаем списки
        lst_prac = df_prac['Наименование'].dropna().tolist()  # список для Прак опыта
        lst_prac = processing_punctuation_end_string(lst_prac, ';\n', '- ', '.')
        lst_prac_control = df_prac['Форма'].dropna().tolist()  # список для Прак опыта
        lst_prac_control = processing_punctuation_end_string(lst_prac_control, ';\n', '- ', '.')

        temp_df = pd.DataFrame(columns=['Наименование', 'Форма'],
                               data=[['Практический опыт',''],[lst_prac, lst_prac_control]])
        control_df = pd.concat([control_df, temp_df], axis=0, ignore_index=True)

        # """
        # Обрабатываем лист Умения
        # """
        # df_skills = pd.read_excel(data_pred_prac, sheet_name=skills, usecols='A:B')
        # df_skills.dropna(inplace=True, thresh=1)  # удаляем пустые строки
        # df_skills.fillna('', inplace=True)
        # df_skills.columns = ['Наименование', 'Форма']
        # # создаем списки
        # lst_skills = df_skills['Наименование'].dropna().tolist()  # список для ОК
        # lst_skills = processing_punctuation_end_string(lst_skills, ';\n', '- ', '.')
        # lst_skills_control = df_skills['Форма'].dropna().tolist()  # список для ОК
        # lst_skills_control = processing_punctuation_end_string(lst_skills_control, ';\n', '- ', '.')
        #
        # temp_df = pd.DataFrame(columns=['Наименование', 'Форма'],
        #                        data=[['Умения',''],[lst_skills, lst_skills_control]])
        # control_df = pd.concat([control_df, temp_df], axis=0, ignore_index=True)

        # """
        # Обрабатываем лист ВПД
        # """
        # df_vpd = pd.read_excel(data_pred_prac,sheet_name=vpd,usecols='A:B')
        # df_vpd.dropna(inplace=True, thresh=1)  # удаляем пустые строки
        # df_vpd.fillna('', inplace=True)
        # df_vpd.columns = ['Наименование','Форма']
        # # создаем списки
        # lst_vpd = df_vpd['Наименование'].dropna().tolist() # список для ОК
        # lst_vpd = processing_punctuation_end_string(lst_vpd, ';\n', '- ', '.')
        # lst_vpd_control = df_vpd['Форма'].dropna().tolist() # список для ОК
        # lst_vpd_control = processing_punctuation_end_string(lst_vpd_control, ';\n', '- ', '.')
        #
        # temp_df = pd.DataFrame(columns=['Наименование','Форма'],
        #                        data=[['Виды деятельности',''],[lst_vpd,lst_vpd_control]])
        # control_df = pd.concat([control_df,temp_df],axis=0,ignore_index=True)

        """
            Обрабатываем лист ФГОС
            """
        df_fgos = pd.read_excel(data_pred_prac, sheet_name=fgos, usecols='A:C')
        df_fgos.dropna(inplace=True, thresh=1)  # удаляем пустые строки
        df_fgos.fillna('Не заполнено',inplace=True)

        doc = DocxTemplate(template_pm)

        # Конвертируем датафрейм с описанием программы в список словарей и добавляем туда нужные элементы
        data_program = df_desc_rp.to_dict('records')
        context = data_program[0]
        context['ФИО_ПР'] = accept_PR # Переменные для ФИО утверждающих  ПР

        context['Лич_результаты'] = df_pers_result.to_dict('records')  # добаввляем лист личностных результатов

        # Единичные переменные

        # context['Квалификация'] = df_desc_rp.at[0,'Квалификация']
        context['Объем_ПП'] = sum_volume_pp
        context['Объем_ПП_прак_под'] = sum_practice_pp

        # Лист Содержание преддипломной практики
        context['ПП_разд_объем'] = df_pp_short.to_dict('records')  # делаем таблицу УП
        context['ПП_содер_объем'] = df_pp_content.to_dict('records')  # делаем таблицу УП

        # #лист МТО
        # context['Оборудование'] = lst_obor
        # context['Инструменты_приспособления'] = lst_tecn
        # context['Средства_обучения'] = lst_educ
        # лист Учебные издания
        context['Основные_источники'] = lst_main_source
        context['Дополнительные_источники'] = lst_slave_source
        context['Интернет_источники'] = lst_inet_source
        # Листы данные ОК,ПК
        context['ОК'] = lst_ok
        context['ПК'] = lst_pk

        context['Прак_опыт'] = lst_prac
        context['ВПД'] = lst_vpd
        context['ФГОС_ВД'] = lst_vd
        context['Контроль_ПК'] = df_pk.to_dict('records')
        context['Контроль'] = control_df.to_dict('records')

        # Переменные ФГОС
        # даты
        context['ФГОС_у_дата'] = df_fgos.at[0,'Дата']
        context['ФГОС_з_дата'] = df_fgos.at[1,'Дата']
        context['Прог_з_дата'] = df_fgos.at[2,'Дата']

        # номера
        context['ФГОС_у_номер'] = df_fgos.at[0,'Номер документа']
        context['ФГОС_з_номер'] = df_fgos.at[1,'Номер документа']
        context['Прог_з_номер'] = df_fgos.at[2,'Номер документа']

        # Создаем документ
        doc.render(context)
        # сохраняем документ
        # название программы
        name_rp = df_desc_rp['Код_специальность_профессия'].tolist()[0]
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        doc.save(f'{end_folder}/РП Преддипломная практика {name_rp[:40]} {current_time}.docx')


    except DiffSheet:
        messagebox.showerror('Диана Создание рабочих программ',
                             f'В таблице не найдены листы {diff_cols},\n'
                             f'В таблице должны быть листы {etalon_cols_lst}\n'
                             f'Возможно вы используете шаблон РП для ООД а не шаблон РП для УД')
    except KeyError as e:
        messagebox.showerror('Диана Создание рабочих программ',
                             f'В таблице не найдена колонка с названием {e.args}!\nПроверьте написание названия колонки')
    except ValueError as e:
        messagebox.showerror('Диана Создание рабочих программ',
                             f'В таблице не найден лист,колонка или значение {e.args}!\nПроверьте написание названий')

    except FileNotFoundError:
        messagebox.showerror('Диана Создание рабочих программ',
                             f'Перенесите файлы которые вы хотите обработать в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам')

    except exceptions.TemplateSyntaxError:
        messagebox.showerror('Диана Создание рабочих программ',
                             f'Ошибка в оформлении вставляемых значений в шаблоне\n'
                             f'Проверьте свой шаблон на наличие следующих ошибок:\n'
                             f'1) Вставляемые значения должны быть оформлены двойными фигурными скобками\n'
                             f'{{{{Вставляемое_значение}}}}\n'
                             f'2) В названии колонки в таблице откуда берутся данные - есть пробелы,цифры,знаки пунктуации и т.п.\n'
                             f'в названии колонки должны быть только буквы и нижнее подчеркивание.\n'
                             f'{{{{Дата_рождения}}}}')
    else:
        messagebox.showinfo('Диана Создание рабочих программ', 'Данные успешно обработаны')



if __name__ == '__main__':
    template_pred_main = 'data/Шаблон автозаполнения преддипломной практики.docx'
    data_pred_main = 'data/Таблица для преддипломной практики.xlsx'
    end_folder_main = 'data'

    create_pred_dip_prac(template_pred_main, data_pred_main, end_folder_main)
    print('Lindy Booth')

